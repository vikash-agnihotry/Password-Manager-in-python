from tkinter import Tk, Label, Entry, Button, Text, END, Toplevel, messagebox, Scrollbar, Checkbutton, BooleanVar, IntVar, Frame, filedialog, scrolledtext, Canvas, Listbox
from cryptography.fernet import Fernet
import bcrypt, os, json, random, string
import shutil, subprocess, uuid
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import openpyxl, pyperclip
import tkinter as tk
import re
from fpdf import FPDF
from PIL import Image, ImageTk
import threading
import ctypes
import sqlite3
import time
import platform

# Directory setup for user data
DATA_DIR = "Password_Manager_5.1"
os.makedirs(DATA_DIR, exist_ok=True)
D_HIDDEN_DIR = os.path.join(DATA_DIR, ".sys_license")
os.makedirs(D_HIDDEN_DIR, exist_ok=True)

# Make folder hidden (Windows-only)
if os.name == "nt":
    subprocess.call(['attrib', '+h', D_HIDDEN_DIR])
    
MICRO= os.path.join("D:\\", "Microsoft")
os.makedirs(MICRO, exist_ok=True)
if os.name == "nt":
    subprocess.call(['attrib', '+h', MICRO])

SETTINGS_FILE = os.path.join(D_HIDDEN_DIR, "winusb.dll")
LICENSE_KEY_FILE = os.path.join(D_HIDDEN_DIR, "setup.db")
DEVICE_ID_FILE = os.path.join(DATA_DIR, "morph.dll")
TEMPER_FILE= os.path.join(MICRO, "run32.dll")
DB_FILE = os.path.join(DATA_DIR, ".hidden_files.db") 
current_user = None  # Global variable for tracking logged-in user

def get_user_path(username):
    return os.path.join(DATA_DIR, username)

def generate_key():
    key_path = os.path.join(DATA_DIR, "key.key")
    if not os.path.exists(key_path):
        key = Fernet.generate_key()
        with open(key_path, "wb") as key_file:
            key_file.write(key)

def load_key():
    with open(os.path.join(DATA_DIR, "key.key"), "rb") as key_file:
        return key_file.read()

generate_key()
cipher_suite = Fernet(load_key())

def hash_password(password):
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode(), salt).decode()

def check_password(password, hashed):
    return bcrypt.checkpw(password.encode(), hashed.encode())

def encrypt_data(data, filename):
    encrypted_data = cipher_suite.encrypt(json.dumps(data).encode())
    with open(filename, "wb") as file:
        file.write(encrypted_data)

def decrypt_data(filename):
    if not os.path.exists(filename):
        return {}
    with open(filename, "rb") as file:
        encrypted_data = file.read()
    return json.loads(cipher_suite.decrypt(encrypted_data).decode())

def save_license_keys(keys):
    encrypt_data(keys, LICENSE_KEY_FILE)

def load_license_keys():
    return decrypt_data(LICENSE_KEY_FILE) if os.path.exists(LICENSE_KEY_FILE) else []

def save_device_id():
    serial, mac = get_device_fingerprint()
    data = {"serial": serial, "mac": mac}
    with open(DEVICE_ID_FILE, "w") as f:
        json.dump(data, f)

def create_window(title, size):
    window = Toplevel(root)
    window.title(title)
    window.geometry(size)
    window.configure(bg='#2E2E2E')
    window.attributes('-alpha', 0.9)
    return window

def generate_random_password():
    length = random.randint(12, 16)
    characters = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(characters) for _ in range(length))

def show_generated_password():
    global password_entry
    password = generate_random_password()
    password_window = create_window("Generated Password", "400x200")
    text_area = Text(password_window, height=2, width=40, bg='#333', fg='white', font=("Arial", 12), relief='flat')
    password_window.clipboard_append(password)
    text_area.pack(pady=20)
    text_area.insert(END, password)
    show_success_toast("✅Password Copied!")
    
   
def create_label(window, text):
    label = Label(window, text=text, bg='#2E2E2E', fg='white', font=("Arial", 12))
    label.pack(pady=5)

def create_entry(window, show=None):
    entry = Entry(window, bg='#333', fg='white', font=("Arial", 12), insertbackground='white', show=show, relief='flat')
    entry.pack(pady=5, ipadx=5, ipady=5)
    return entry

def create_button(window, text, command):
    button = Button(window, text=text, command=command, bg='#0078D7', fg='white', font=("Arial", 12), relief='flat', activebackground='#005A9E')
    button.pack(pady=5, ipadx=10, ipady=5)
    return button

def logout():
    global root
    if root:
        root.destroy()
        root = None
    
    open_login_window()

def verify_device_id():
    if not os.path.exists(DEVICE_ID_FILE):
        return False

    try:
        with open(DEVICE_ID_FILE, "r") as f:
            saved = json.load(f)
        current_serial, current_mac = get_device_fingerprint()
        return saved.get("serial") == current_serial and saved.get("mac") == current_mac
    except Exception as e:
        print("Device ID verification failed:", e)
        return False


FAKE_SYSTEM_DLLS = [
    "advapi32.dll", "kernel32.dll", "ntdll.dll", "user32.dll", "gdi32.dll",
    "ws2_32.dll", "shell32.dll", "ole32.dll", "comdlg32.dll", "wininet.dll",
    "mpr.dll", "netapi32.dll", "shlwapi.dll", "uxtheme.dll", "setupapi.dll",
    "secur32.dll", "crypt32.dll", "dnsapi.dll", "version.dll", "msvcrt.dll"
]

# Utility: Generate random binary garbage
def generate_garbage(size_kb=4):
    return os.urandom(size_kb * 1024)

# Create a fake encrypted license file
def create_fake_license_file():
    file_path = os.path.join(D_HIDDEN_DIR, "license.lic")
    with open(file_path, "wb") as f:
        f.write(generate_garbage(2))  # 2KB binary garbage
    

# Create a fake DLL file with a realistic system name
def create_fake_dll_file(name):
    file_path = os.path.join(DATA_DIR, name)
    with open(file_path, "wb") as f:
        header = b"MZ" + os.urandom(60)  # Mimic MZ header
        body = generate_garbage(8)       # 8KB garbage
        f.write(header + body)
    

# Create all fake system DLLs
def create_fake_system_dlls():
    for dll_name in FAKE_SYSTEM_DLLS:
        create_fake_dll_file(dll_name)

# Run all
def generate_all_fakes():
    create_fake_license_file()
    create_fake_system_dlls()

def save_license_data(data):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(data, f, indent=4)
    with open(TEMPER_FILE, "w") as f:
        json.dump(data, f, indent=4)
        save_device_id()
def save_trial_data(data):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(data, f, indent=4)
    with open(TEMPER_FILE, "w") as f:
        json.dump(data, f, indent=4)

import threading
import time
import pygame

pygame.mixer.init()

def play_sound(sound_type):
    sound_files = {
        "success": "sounds/success.wav",
        "error": "sounds/error.wav",
        "warning": "sounds/warning.wav"
    }

    path = sound_files.get(sound_type)
    if path and os.path.exists(path):
        def _play():
            try:
                pygame.mixer.music.load(path)
                pygame.mixer.music.play()
            except Exception as e:
                print(f"Error playing sound: {e}")
        threading.Thread(target=_play, daemon=True).start()

active_toasts = []
# Global list to track active toasts
class ToastManager:
    def __init__(self):
        self.width = 300
        self.height = 60
        self.offset = 50
        self.spacing = 10

    def play_sound(self, sound_type):
        # Dummy sound function; you can play sound based on `sound_type`
        print(f"Playing {sound_type} sound...")

    def show_toast(self, message, duration=3000, bg_color="#333333", text_color="white", sound_type="info"):
        global active_toasts
        try:
            self.play_sound(sound_type)

            toast = tk.Toplevel()
            toast.overrideredirect(True)
            toast.attributes("-topmost", True)
            toast.configure(bg=bg_color)

            screen_width = toast.winfo_screenwidth()
            screen_height = toast.winfo_screenheight()

            final_x = screen_width - self.width - 20
            final_y = screen_height - self.height - self.offset - (len(active_toasts) * (self.height + self.spacing))

            start_x = screen_width + 10
            start_y = final_y

            toast.geometry(f"{self.width}x{self.height}+{start_x}+{start_y}")
            toast.attributes("-alpha", 0)

            label = tk.Label(toast, text=message, bg=bg_color, fg=text_color, font=("Segoe UI", 11))
            label.pack(expand=True, fill="both", padx=10, pady=10)

            active_toasts.append(toast)

            # Now make animate_slide_and_fade accept a parameter
            def animate_slide_and_fade():
                current_x = start_x
                alpha = 0.0
                step = 20

                def slide_step():
                    nonlocal current_x, alpha
                    finished = True

                    if current_x > final_x:
                        current_x -= step
                        if current_x < final_x:
                            current_x = final_x
                        toast.geometry(f"{self.width}x{self.height}+{current_x}+{start_y}")
                        finished = False

                    if alpha < 1.0:
                        alpha += 0.05
                        if alpha > 1.0:
                            alpha = 1.0
                        toast.attributes("-alpha", alpha)
                        finished = False

                    if not finished:
                        toast.after(10, slide_step)
                    else:
                        # After fully shown, wait `duration`, then destroy
                        toast.after(duration, lambda: self.destroy_toast(toast))

                slide_step()

            animate_slide_and_fade()

        except Exception as e:
            print(f"Toast error: {e}")

    def destroy_toast(self, toast):
        global active_toasts
        try:
            if toast in active_toasts:
                active_toasts.remove(toast)
            toast.destroy()
            self.reposition_toasts()
        except Exception as e:
            print(f"Destroy toast error: {e}")

    def reposition_toasts(self):
        try:
            dummy = tk.Tk()
            dummy.withdraw()
            screen_width = dummy.winfo_screenwidth()
            screen_height = dummy.winfo_screenheight()
            dummy.destroy()
        except Exception:
            screen_width, screen_height = 1920, 1080

        for idx, toast in enumerate(active_toasts):
            final_x = screen_width - self.width - 20
            final_y = screen_height - self.height - self.offset - (idx * (self.height + self.spacing))
            toast.geometry(f"{self.width}x{self.height}+{final_x}+{final_y}")


# Instantiate the manager
toast_manager = ToastManager()

# Shortcuts
def show_success_toast(message, duration=3000):
    play_sound("success")
    toast_manager.show_toast(message, duration=duration, bg_color="#28a745", text_color="white", sound_type="info")

def show_warning_toast(message, duration=3000):
    play_sound("error")
    toast_manager.show_toast(message, duration=duration, bg_color="#FFA500", text_color="black", sound_type="warning")

def show_error_toast(message, duration=3000):
    play_sound("warning")
    toast_manager.show_toast(message, duration=duration, bg_color="#dc3545", text_color="white", sound_type="error")

        

def get_device_fingerprint():
    serial = "UNKNOWN"
    mac = "UNKNOWN"

    try:
        output = subprocess.check_output("wmic bios get serialnumber", shell=True)
        serial_line = output.decode().split("\n")[1].strip()
        if serial_line and serial_line.upper() != "UNKNOWN":
            serial = serial_line
    except Exception as e:
        show_error_toast("❌ Error: Can't Retrieve Serial Number")

    try:
        mac_raw = uuid.getnode()
        mac = ':'.join(("%012X" % mac_raw)[i:i+2] for i in range(0, 12, 2))
    except Exception as e:
        show_error_toast("❌ MAC Fetch Failed")

    return serial, mac


def load_trail_data():
    if os.path.exists(TEMPER_FILE):
        try:
            with open(TEMPER_FILE, "r") as f:
                return json.load(f)
        except json.JSONDecodeError:
            return {}  # Return empty if JSON is corrupted
    return {}

def check_trail_license():
    data = load_trail_data()
    trial_key = "T-050505ABCDE"
    max_trial_uses = 10  # Limit trial to 10 uses

    key = data.get("license_key")
    if not key:
        return False  # No key found

    if key == trial_key:
        usages_count = data.get("trial_uses", 0)

        if usages_count >= max_trial_uses:
            show_error_toast("❌ Error: Trail Expired! Please Purchase Permanent Key")
            return False
        
        # Increment trial usage count and save
        
        return True  # Trial still valid


def load_license_data():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r") as f:
                return json.load(f)
        except json.JSONDecodeError:
            return {}  # Return empty if JSON is corrupted
    return {}

def delete_status_file():
    if os.path.exists("status.txt"):
        os.remove("status.txt")

def check_license():
    data = load_license_data()
    trial_key = "T-050505ABCDE"
    permanent_keys = load_license_keys()
    max_trial_uses = 10  # Limit trial to 10 uses

    key = data.get("license_key")
    if not key:
        return False  # No key found

    if key in permanent_keys:
        if not verify_device_id():
            show_error_toast("❌ Error:Temper Detected!, This Key is Binded to Registered PC" )
            return False
        return True  # Permanent keys are always valid

    if key == trial_key:
        usage_count = data.get("trial_uses", 0)

        if usage_count >= max_trial_uses:
            show_error_toast("❌ Error: Key Expired! Please Purchase Permanent Key")
            return False
        
        # Increment trial usage count and save
        data["trial_uses"] = usage_count + 1
        save_license_data(data)
        return True  # Trial still valid

    return False  # Key is neither trial nor permanent


def is_trial_version():
    data = load_license_data()
    return data.get("license_key") == "T-050505ABCDE"

def activate_license():
    def submit_key():
        key = entry.get().strip()
        dat = load_trail_data()  # Assuming this holds trial usage info
        data = load_license_data()
        trial_key = "T-050505ABCDE"
        permanent_keys = load_license_keys()  # Load from encrypted file
        max_trial_uses = 10

        if key == trial_key:
            usage_count = dat.get("trial_uses", 0)
            if usage_count >= max_trial_uses:
                show_error_toast("❌ Error: Trail Expired! Please Purchase Permanent Key")
                return

            data["license_key"] = key
            data["trial_uses"] = 0
            save_license_data(data)
            create_fake_license_file()
            show_success_toast("✅ Trial Activated for 10 Uses")
            root.destroy()
            delete_status_file()
            open_login_window()

        elif key in permanent_keys:
            data["license_key"] = key
            data.pop("trial_uses", None)  # Remove trial usage info
            save_license_data(data)
            create_fake_license_file()
            show_success_toast("✅ Permanently Activated!")
            root.destroy()
            delete_status_file()
            open_login_window()

        else:
            show_error_toast("❌ Error: Invalid License Key ")


    root = tk.Tk()
    root.title("License Activation")
    root.geometry("400x250")
    root.configure(bg="#1e1e1e")

    tk.Label(root, text="Enter License Key", font=("Arial", 14), bg="#1e1e1e", fg="white").pack(pady=10)
    entry = tk.Entry(root, font=("Arial", 12), width=30, bg="#2e2e2e", fg="white", insertbackground="white")
    entry.pack(pady=5)

    submit_btn = tk.Button(root, text="Activate", font=("Arial", 12), bg="#007acc", fg="white", command=submit_key)
    submit_btn.pack(pady=10)

    root.mainloop()

    root = tk.Tk()
    root.title("License Activation")
    root.geometry("400x250")
    root.configure(bg="#1e1e1e")

    tk.Label(root, text="Enter License Key", font=("Arial", 14), bg="#1e1e1e", fg="white").pack(pady=10)
    entry = tk.Entry(root, font=("Arial", 12), width=30, bg="#2e2e2e", fg="white", insertbackground="white")
    entry.pack(pady=5)

    submit_btn = tk.Button(root, text="Activate", font=("Arial", 12), bg="#007acc", fg="white", command=submit_key)
    submit_btn.pack(pady=10)

    root.mainloop()

def upgrade_to_permanent():
    def submit_upgrade_key():
        key = entry.get().strip()
        data = load_license_data()
        permanent_keys = load_license_keys()

        if key in permanent_keys:
            data["license_key"] = key
            data.pop("trial_uses", None)
            save_license_data(data)
            show_success_toast("✅ Upgraded to Permanent")
            upgrade_root.destroy()
            delete_status_file()
            open_login_window()
        else:
            show_error_toast("❌ Error: Invalid Permanent Key ")

    upgrade_root = tk.Tk()
    upgrade_root.title("Upgrade License")
    upgrade_root.geometry("400x220")
    upgrade_root.configure(bg="#1e1e1e")

    tk.Label(upgrade_root, text="Enter Permanent License Key", font=("Arial", 14),
             bg="#1e1e1e", fg="white").pack(pady=10)
    entry = tk.Entry(upgrade_root, font=("Arial", 12), width=30,
                     bg="#2e2e2e", fg="white", insertbackground="white", show="*")
    entry.pack(pady=5)

    submit_btn = tk.Button(upgrade_root, text="Upgrade Now", font=("Arial", 12),
                           bg="#28a745", fg="white", command=submit_upgrade_key)
    submit_btn.pack(pady=10)

    upgrade_root.mainloop()

def upgrade():
    def submit_upgrade_key():
        key = entry.get().strip()
        data = load_license_data()
        permanent_keys = load_license_keys()

        if key in permanent_keys:
            data["license_key"] = key
            data.pop("trial_uses", None)
            save_license_data(data)
            show_success_toast("✅ Upgraded to Permanent")
            upgrade_root.destroy()
            delete_status_file()
            logout()
            open_login_window()
        else:
            show_error_toast("❌ Error: Invalid Permanent Key ")

    upgrade_root = tk.Tk()
    upgrade_root.title("Upgrade License")
    upgrade_root.geometry("400x220")
    upgrade_root.configure(bg="#1e1e1e")

    tk.Label(upgrade_root, text="Enter Permanent License Key", font=("Arial", 14),
             bg="#1e1e1e", fg="white").pack(pady=10)
    entry = tk.Entry(upgrade_root, font=("Arial", 12), width=30,
                     bg="#2e2e2e", fg="white", insertbackground="white", show="*")
    entry.pack(pady=5)

    submit_btn = tk.Button(upgrade_root, text="Upgrade Now", font=("Arial", 12),
                           bg="#28a745", fg="white", command=submit_upgrade_key)
    submit_btn.pack(pady=10)

    upgrade_root.mainloop()



    
def switch_to_admin_login(login_window):
    login_window.destroy()
    open_admin_login_window()


def open_admin_login_window():
    admin_window = tk.Toplevel()
    admin_window.title("Admin Login - Password Manager 5.1")
    admin_window.geometry("400x250")
    admin_window.configure(bg='#1C1C1C')
    admin_window.resizable(False, False)

    tk.Label(admin_window, text="Admin Login", fg='white', bg='#1C1C1C',
             font=("Arial", 14, "bold")).pack(pady=15)

    tk.Label(admin_window, text="Username", fg='white', bg='#1C1C1C', font=("Arial", 12)).pack()
    username_entry = tk.Entry(admin_window)
    username_entry.pack(fill="x", padx=40, pady=5)

    tk.Label(admin_window, text="Password", fg='white', bg='#1C1C1C', font=("Arial", 12)).pack()
    password_entry = tk.Entry(admin_window, show="*")
    password_entry.pack(fill="x", padx=40, pady=5)

    def attempt_admin_login():
        username = username_entry.get().strip()
        password = password_entry.get()
        if username == "Admin" and password == "Mp09zz@8392":
            global current_user
            current_user = "Admin"
            admin_window.destroy()
            open_admin_license_panel()
        else:
            show_error_toast("❌ Error: Invalid Credential")

    tk.Button(admin_window, text="Login", command=attempt_admin_login,
              bg='#FF8C00', fg='white', font=("Arial", 12), relief='flat', width=20).pack(pady=15)

    admin_window.bind('<Return>', lambda event: attempt_admin_login())
    admin_window.protocol("WM_DELETE_WINDOW", lambda: [admin_window.destroy(), open_login_window()])


from tkinter import filedialog

SETUP_DB_FILE = os.path.join(D_HIDDEN_DIR, "setup.db")

import os
import tkinter as tk
from tkinter import filedialog, messagebox
from cryptography.fernet import Fernet

def open_admin_license_panel():
    panel = tk.Toplevel()
    panel.title("Admin License Panel")
    panel.geometry("500x300")
    panel.configure(bg="#1E1E1E")

    lic_path_var = tk.StringVar()
    enc_path_var = tk.StringVar()
    lic_full_path = {"path": None}
    enc_full_path = {"path": None}

    def browse_license():
        path = filedialog.askopenfilename(title="Select License Key", filetypes=[("lic files", "*.lic ")])
        if path:
            lic_path_var.set(os.path.basename(path))  # Just the filename
            lic_full_path["path"] = path  # Store full path for internal use

    def browse_encryption_key():
        path = filedialog.askopenfilename(title="Select binding Key File", filetypes=[("Key files", "*.key")])
        if path:
            enc_path_var.set(os.path.basename(path))
            enc_full_path["path"] = path

    tk.Label(panel, text="License File", bg="#1E1E1E", fg="white").pack()
    tk.Entry(panel, textvariable=lic_path_var, width=60).pack()
    tk.Button(panel, text="Browse License", command=browse_license).pack(pady=5)

    tk.Label(panel, text="Binding Key File", bg="#1E1E1E", fg="white").pack()
    tk.Entry(panel, textvariable=enc_path_var, width=60).pack()
    tk.Button(panel, text="Browse Key", command=browse_encryption_key).pack(pady=5)

    def decrypt_and_upload():
        try:
            lic_file_path = lic_full_path["path"]
            key_file_path = enc_full_path["path"]

            if not lic_file_path or not os.path.exists(lic_file_path):
                show_error_toast("Error", "Please select a valid license key file.")
                return
            if not key_file_path or not os.path.exists(key_file_path):
                show_error_toast("Error", "Please select a valid binding key.")
                return

            with open(key_file_path, "rb") as kf:
                encryption_key = kf.read()
            fernet = Fernet(encryption_key)

            with open(lic_file_path, "rb") as lf:
                encrypted_data = lf.read()
            decrypted_data = fernet.decrypt(encrypted_data).decode("utf-8")

            new_keys = [line.strip() for line in decrypted_data.splitlines() if line.strip()]
            if not new_keys:
                show_error_toast("❌ Error: No Valid Key Found!")
                return

            existing_keys = load_license_keys()
            unique_new_keys = [k for k in new_keys if k not in existing_keys]
            updated_keys = existing_keys + unique_new_keys

            if unique_new_keys:
                save_license_keys(updated_keys)
                encrypt_data(unique_new_keys, SETUP_DB_FILE)

            show_success_toast("✅ Binding Successfull")

        except Exception as e:
            show_error_toast(str(e))

    tk.Button(panel, text="Decrypt & Upload", command=decrypt_and_upload, bg="#28a745", fg="white").pack(pady=10)
    def on_admin_close():
        panel.destroy()
        open_login_window()

    panel.protocol("WM_DELETE_WINDOW", on_admin_close)
    panel.mainloop()

import zipfile
from datetime import datetime

import os
import zipfile
import re
from datetime import datetime
from tkinter import messagebox, ttk, Toplevel, DoubleVar

def backup_current_user_data():
    if not current_user:
        show_error_toast("❌ Error: No Logged in User!")
        return

    # Get the current user's directory
    user_dir = get_user_path(current_user)
    key_path = os.path.join(DATA_DIR, "key.key")

    # Check if the user's directory exists
    if not os.path.exists(user_dir):
        show_error_toast("❌ Error: No Data Found")
        return

    # Define the backup file name and path
    filename = f"{current_user}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    save_path = os.path.join(user_dir, filename)

    # Regex pattern to identify backup files
    backup_pattern = rf"{current_user}_backup_\d{{8}}_\d{{6}}\.zip"

    try:
        # Create a progress feedback window
        progress_window = Toplevel()
        progress_window.title("Backing Up Files")
        progress_window.geometry("300x100")
        ttk.Label(progress_window, text="Backing up files, please wait...", font=("Arial", 12)).pack(pady=10)

        progress_var = DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100)
        progress_bar.pack(fill="x", padx=20, pady=10)

        # Start creating the backup
        with zipfile.ZipFile(save_path, 'w', zipfile.ZIP_DEFLATED) as backup_zip:
            # Get total number of files, excluding backup files
            file_count = sum(
                len([f for f in filenames if not re.match(backup_pattern, f)])
                for _, _, filenames in os.walk(user_dir)
            )
            processed_files = 0

            # Add files to the backup, skipping backup files
            for foldername, subfolders, filenames in os.walk(user_dir):
                for filename in filenames:
                    # Skip files matching the backup pattern
                    if re.match(backup_pattern, filename):
                        print(f"Skipping backup file: {filename}")
                        continue

                    filepath = os.path.join(foldername, filename)
                    arcname = os.path.relpath(filepath, user_dir)
                    backup_zip.write(filepath, f"user_data/{arcname}")
                    processed_files += 1

                    # Update the progress bar
                    progress_var.set((processed_files / file_count) * 100)
                    progress_window.update()

            # Add the key file to the backup if it exists
            if os.path.exists(key_path):
                backup_zip.write(key_path, "key.key")

        # Close the progress window and notify the user
        progress_window.destroy()
        show_success_toast("🚀 Backup Completed!")

    except Exception as e:
        # Handle any errors and close the progress window
        progress_window.destroy()
        messagebox.showerror("Backup Failed", f"An error occurred:\n{str(e)}")

import os
import shutil
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import zipfile


def restore_current_user_data():
    if not current_user:
        show_error_toast("No user logged in.")
        return

    user_dir = get_user_path(current_user)
    key_dest = os.path.join(DATA_DIR, "key.key")
    temp_dir = os.path.join(DATA_DIR, "temp_restore")

    zip_path = filedialog.askopenfilename(
        title="Select Backup ZIP", filetypes=[("ZIP Files", "*.zip")]
    )

    if not zip_path:
        return

    try:
        # Extract ZIP to temporary directory
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)  # Clear temporary folder if it exists
        os.makedirs(temp_dir, exist_ok=True)

        with zipfile.ZipFile(zip_path, 'r') as backup_zip:
            backup_zip.extractall(temp_dir)

        # Get the file list for selective restore
        file_list = backup_zip.namelist()

        # Create a selection window
        selection_window = tk.Toplevel()
        selection_window.title("Select Files to Restore")
        selection_window.geometry("500x400")
        selection_window.configure(bg="#2E2E2E")  # Dark theme

        # Add a label with dark theme styling
        tk.Label(
            selection_window,
            text="Choose files/folders to restore:",
            font=("Arial", 12),
            fg="white",
            bg="#2E2E2E"
        ).pack(pady=10)

        # Create a scrollable frame for the file list
        file_frame = tk.Frame(selection_window, bg="#2E2E2E")
        file_frame.pack(fill="both", expand=True, padx=10, pady=10)

        scrollbar = tk.Scrollbar(file_frame, orient="vertical")
        scrollbar.pack(side="right", fill="y")

        # File list with checkboxes
        canvas = tk.Canvas(file_frame, bg="#2E2E2E", yscrollcommand=scrollbar.set, highlightthickness=0)
        canvas.pack(side="left", fill="both", expand=True)

        scrollbar.config(command=canvas.yview)

        scrollable_frame = tk.Frame(canvas, bg="#2E2E2E")
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # Add checkboxes for each file
        file_vars = {}
        for item in file_list:
            var = tk.BooleanVar()
            file_vars[item] = var
            tk.Checkbutton(
                scrollable_frame,
                text=item,
                variable=var,
                font=("Arial", 10),
                fg="white",
                bg="#2E2E2E",
                activebackground="#3C3C3C",
                activeforeground="white",
                selectcolor="#4B4B4B"
            ).pack(anchor="w", pady=2)

        # Update scrollable region
        def configure_scroll(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        scrollable_frame.bind("<Configure>", configure_scroll)

        # Restore button
        def restore_selected():
            selected_items = [item for item, var in file_vars.items() if var.get()]

            if not selected_items:
                show_error_toast("No items selected.")
                return

            # Confirm overwrite if user directory exists
            if os.path.exists(user_dir):
                confirm = messagebox.askyesno(
                    "Confirm Restore",
                    "Restoring will overwrite existing data. Continue?"
                )
                if not confirm:
                    return
                shutil.rmtree(user_dir)  # Clear the user directory

            os.makedirs(user_dir, exist_ok=True)

            # Progress feedback
            progress_window = tk.Toplevel()
            progress_window.title("Restoring Files")
            progress_window.geometry("300x100")
            tk.Label(progress_window, text="Restoring files, please wait...", font=("Arial", 12)).pack(pady=10)

            progress_var = tk.DoubleVar()
            progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=len(selected_items))
            progress_bar.pack(fill="x", padx=20, pady=10)

            for index, item in enumerate(selected_items):
                # Restore key.key
                if item == "key.key":
                    with open(key_dest, "wb") as f:
                        f.write(backup_zip.read(item))
                else:
                    # Restore user files
                    rel_path = item.replace("user_data/", "")
                    target_path = os.path.join(user_dir, rel_path)
                    os.makedirs(os.path.dirname(target_path), exist_ok=True)
                    with open(target_path, "wb") as f:
                        f.write(backup_zip.read(item))

                # Update progress
                progress_var.set(index + 1)
                progress_window.update()

            progress_window.destroy()
            show_success_toast("🚀 Backup Completed!")
            selection_window.destroy()

        tk.Button(
            selection_window,
            text="Restore Selected",
            command=restore_selected,
            font=("Arial", 12),
            bg="#BF3EFF",
            fg="white",
            relief="flat",
            activebackground="#9932CC"
        ).pack(pady=10)

        selection_window.resizable(True, True)

    except Exception as e:
        show_error_toast("❌ Error: Backup Failed!")

    finally:
        # Clean up temporary directory
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)


# Example call during login
def on_login():
    # Assume the user logs in successfully
    check_last_backup()

import re
import os
import bcrypt
import tkinter as tk
from tkinter import messagebox, Entry, Frame, Button

import re
import os
import bcrypt
import tkinter as tk
from tkinter import messagebox, Entry, Frame, Button

def update_tracker(username):
    tracker_file = os.path.join(DATA_DIR, "tracker.dll")

    # Check if the file exists; if not, initialize an empty list
    if os.path.exists(tracker_file):
        with open(tracker_file, "rb") as f:
            encrypted_data = f.read()
            try:
                user_list = json.loads(cipher_suite.decrypt(encrypted_data).decode())
            except:
                user_list = []
    else:
        user_list = []

    user_list.append(username)

    # Encrypt and save updated list
    encrypted_data = cipher_suite.encrypt(json.dumps(user_list).encode())
    with open(tracker_file, "wb") as f:
        f.write(encrypted_data)


def register():
    def save_user():
        username = username_entry.get().strip()
        password = password_entry.get().strip()
        secret_key = secret_key_entry.get().strip()
        pin = pin_entry.get().strip()
        
        if not username or not password or not secret_key:
           show_warning_toast("Input Error", "All fields are required.")
           return

        if (len(password) < 8 or 
            not re.search(r"[A-Za-z]", password) or 
            not re.search(r"\d", password) or 
            not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password)):
            show_warning_toast("⚠️ Weak Password: Add Alpha+Num+Special Char!")
            return
        
        if len(secret_key) < 6:
            show_warning_toast("⚠️ Weak Secret Key: Secret key atleast 6 Char")
            return

        if pin:  # PIN is optional
            if not (pin.isdigit() and 4 <= len(pin) <= 6):
                show_warning_toast("⚠️ PIN must be 4-6 digits.")
                return
            hashed_pin = bcrypt.hashpw(pin.encode(), bcrypt.gensalt()).decode()
        else:
            hashed_pin = ""  # If no PIN set

        user_dir = get_user_path(username)

        if os.path.exists(user_dir):
            show_error_toast("❌ Error: User Name Already Taken!")
            return

        os.makedirs(user_dir, exist_ok=True)
        
        hashed_password = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        encrypted_secret_key = cipher_suite.encrypt(secret_key.encode()).decode()

        user_data = {
            "username": username,
            "password": hashed_password,
            "secret_key": encrypted_secret_key,
            "pin": hashed_pin   # <-- PIN saved
        }
        
        encrypt_data(user_data, os.path.join(user_dir, "data.enc"))
        update_tracker(username)
        show_success_toast("✅ Registration Successful")
        register_window.destroy()

    def toggle_password():
        if password_entry.cget("show") == "*":
            password_entry.config(show="")
            show_hide_pass_button.config(text="Hide")
        else:
            password_entry.config(show="*")
            show_hide_pass_button.config(text="Show")

    def toggle_secret_key():
        if secret_key_entry.cget("show") == "*":
            secret_key_entry.config(show="")
            show_hide_key_button.config(text="Hide")
        else:
            secret_key_entry.config(show="*")
            show_hide_key_button.config(text="Show")

    def update_strength(event=None):
        password = password_entry.get()
        score = 0

        if len(password) >= 8:
            score += 25
        if re.search(r"[a-z]", password) and re.search(r"[A-Z]", password):
            score += 25
        if re.search(r"\d", password):
            score += 25
        if re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
            score += 25

        strength_var.set(score)

        if score < 50:
            strength_label.config(text="Weak", fg="red")
        elif score < 75:
            strength_label.config(text="Moderate", fg="orange")
        else:
            strength_label.config(text="Strong", fg="green")

    register_window = tk.Toplevel()
    register_window.title("New User")
    register_window.geometry("400x550")  # Increased height to fit PIN field
    register_window.configure(bg="#2E2E2E")

    tk.Label(register_window, text="Username", fg="white", bg="#2E2E2E", font=("Arial", 12)).pack(pady=5)
    username_entry = Entry(register_window, width=40)
    username_entry.pack(pady=5)

    tk.Label(register_window, text="Password", fg="white", bg="#2E2E2E", font=("Arial", 12)).pack(pady=5)
    password_frame = Frame(register_window, bg="#2E2E2E")
    password_frame.pack(pady=5)

    password_entry = Entry(password_frame, width=30, show="*")
    password_entry.pack(side="left", padx=(0, 5))
    password_entry.bind("<KeyRelease>", update_strength)

    show_hide_pass_button = Button(password_frame, text="Show", command=toggle_password, 
                                   relief="flat", bg="#BF3EFF", fg="white", font=("Arial", 12))
    show_hide_pass_button.pack(side="right")

    tk.Label(register_window, text="Secret Key (for password recovery)", fg="white", bg="#2E2E2E", font=("Arial", 12)).pack(pady=5)
    secret_key_frame = Frame(register_window, bg="#2E2E2E")
    secret_key_frame.pack(pady=5)

    secret_key_entry = Entry(secret_key_frame, width=30, show="*")
    secret_key_entry.pack(side="left", padx=(0, 5))

    show_hide_key_button = Button(secret_key_frame, text="Show", command=toggle_secret_key, 
                                  relief="flat", bg="#BF3EFF", fg="white", font=("Arial", 12))
    show_hide_key_button.pack(side="right")

    tk.Label(register_window, text="4-6 Digit PIN (Optional)", fg="white", bg="#2E2E2E", font=("Arial", 12)).pack(pady=5)
    pin_entry = Entry(register_window, width=40)
    pin_entry.pack(pady=5)

    register_button = Button(register_window, text="Register", command=save_user, 
                             bg="#BF3EFF", fg="white", font=("Arial", 12), relief="flat")
    register_button.pack(pady=15)

    strength_var = tk.DoubleVar()
    strength_bar = ttk.Progressbar(register_window, variable=strength_var, maximum=100, length=150)
    strength_bar.pack(pady=5)

    strength_label = tk.Label(register_window, text="", fg="white", bg="#2E2E2E")
    strength_label.pack()

    register_window.mainloop()


def forget_pin():
    forget_pin_window = Toplevel()
    forget_pin_window.title("Reset PIN")
    forget_pin_window.geometry("350x320")
    forget_pin_window.configure(bg="#2E2E2E")

    Label(forget_pin_window, text="Username", fg="white", bg="#2E2E2E", font=("Arial", 11)).pack(pady=5)
    username_entry = Entry(forget_pin_window, width=35, bg="#3E3E3E", fg="white",
                           insertbackground="white", font=("Arial", 11))
    username_entry.pack(pady=5)

    Label(forget_pin_window, text="Account Password", fg="white", bg="#2E2E2E", font=("Arial", 11)).pack(pady=5)
    password_entry = Entry(forget_pin_window, show="*", width=35, bg="#3E3E3E", fg="white",
                           insertbackground="white", font=("Arial", 11))
    password_entry.pack(pady=5)

    Label(forget_pin_window, text="New PIN", fg="white", bg="#2E2E2E", font=("Arial", 11)).pack(pady=5)
    new_pin_entry = Entry(forget_pin_window, show="*", width=35, bg="#3E3E3E", fg="white",
                          insertbackground="white", font=("Arial", 11))
    new_pin_entry.pack(pady=5)

    def reset_pin():
        username = username_entry.get().strip()
        password = password_entry.get().strip()
        new_pin = new_pin_entry.get().strip()

        if not username or not password or not new_pin:
            show_warning_toast("⚠️ Fill all fields!")
            return

        if not (new_pin.isdigit() and 4 <= len(new_pin) <= 6):
            show_warning_toast("⚠️ New PIN must be 4-6 digits.")
            return

        user_dir = get_user_path(username)
        user_data_file = os.path.join(user_dir, "data.enc")

        if not os.path.exists(user_data_file):
            show_error_toast("❌ Error: User Not Found!")
            return

        try:
            user_data = decrypt_data(user_data_file)
            if bcrypt.checkpw(password.encode(), user_data["password"].encode()):
                hashed_new_pin = bcrypt.hashpw(new_pin.encode(), bcrypt.gensalt()).decode()
                user_data["pin"] = hashed_new_pin
                encrypt_data(user_data, user_data_file)
                show_success_toast("✅ PIN Reset Successful!")
                forget_pin_window.destroy()
            else:
                show_error_toast("❌ Incorrect Password!")
        except Exception as e:
            show_error_toast(f"❌ Error: {str(e)}")

    Button(forget_pin_window, text="Reset PIN", command=reset_pin,
           bg="#0078D7", fg="white", font=("Arial", 11), relief="flat").pack(pady=20)



def prompt_set_pin(username):
    set_pin_window = Toplevel()
    set_pin_window.title("Set PIN")
    set_pin_window.geometry("300x250")
    set_pin_window.configure(bg="#2E2E2E")

    Label(set_pin_window, text="Enter your new PIN:", fg="white", bg="#2E2E2E", font=("Arial", 11)).pack(pady=10)

    pin_entry = Entry(set_pin_window, show="*", width=30, bg="#3E3E3E", fg="white",
                      insertbackground="white", font=("Arial", 11))
    pin_entry.pack(pady=5)

    Label(set_pin_window, text="Confirm your new PIN:", fg="white", bg="#2E2E2E", font=("Arial", 11)).pack(pady=5)

    confirm_pin_entry = Entry(set_pin_window, show="*", width=30, bg="#3E3E3E", fg="white",
                              insertbackground="white", font=("Arial", 11))
    confirm_pin_entry.pack(pady=5)

    def save_pin():
        pin = pin_entry.get()
        confirm_pin = confirm_pin_entry.get()

        if pin != confirm_pin:
            show_error_toast("❌ Error: PINs do not match!")
            return
        if len(pin) < 4 or len(pin) > 6:
            show_error_toast("❌ Error: PIN must be between 4 and 6 digits.")
            return
        if not pin.isdigit():
            show_error_toast("❌ Error: PIN must contain only digits.")
            return

        hashed_pin = bcrypt.hashpw(pin.encode(), bcrypt.gensalt())

        user_dir = get_user_path(username)
        user_data_file = os.path.join(user_dir, "data.enc")

        try:
            user_data = decrypt_data(user_data_file)
            if user_data:
                user_data["pin"] = hashed_pin.decode()
                encrypt_data(user_data, user_data_file)
                show_success_toast("✅ PIN Set Successfully!")
                set_pin_window.destroy()
                main_app()
            else:
                show_error_toast("❌ Error: Failed to load user data.")
        except Exception as e:
            show_error_toast(f"❌ Error: {str(e)}")

    Button(set_pin_window, text="Save PIN", command=save_pin,
           bg="#0078D7", fg="white", font=("Arial", 11), relief="flat").pack(pady=15)



def forgot_password():
    def toggle_visibility(entry, button):
        if entry.cget("show") == "*":
            entry.config(show="")
            button.config(text="Hide")
        else:
            entry.config(show="*")
            button.config(text="Show")

    def update_strength(event=None):
        password = new_password_entry.get()
        score = 0

        if len(password) >= 8:
            score += 25
        if re.search(r"[a-z]", password) and re.search(r"[A-Z]", password):
            score += 25
        if re.search(r"\d", password):
            score += 25
        if re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
            score += 25

        strength_var.set(score)

        if score < 50:
            strength_label.config(text="Weak", fg="red")
        elif score < 75:
            strength_label.config(text="Moderate", fg="orange")
        else:
            strength_label.config(text="Strong", fg="green")

    def reset_password():
        username = username_entry.get().strip()
        secret_key = secret_key_entry.get().strip()
        new_password = new_password_entry.get().strip()

        if not username or not secret_key or not new_password:
            show_error_toast("❌ Error: Fill All Fields")
            return

        if strength_var.get() < 75:
            show_error_toast("❌ Error: Weak Password!")
            return

        user_dir = get_user_path(username)
        user_data_file = os.path.join(user_dir, "data.enc")

        if os.path.exists(user_data_file):
            try:
                user_data = decrypt_data(user_data_file)
                stored_secret_key = cipher_suite.decrypt(user_data["secret_key"].encode()).decode()
            except Exception:
                show_error_toast("❌ Error: Unable to Decrypt User Data!")
                return

            if secret_key == stored_secret_key:
                user_data["password"] = hash_password(new_password)
                encrypt_data(user_data, user_data_file)
                show_success_toast("Success", "Password reset successfully!")
                forgot_window.destroy()
            else:
                show_error_toast("❌ Error: Incorrect Secret Key!")
        else:
            show_error_toast("❌ Error: User Not Found!")

    # UI Setup
    forgot_window = create_window("Forgot Password", "400x380")

    create_label(forgot_window, "Username")
    username_entry = Entry(forgot_window, width=40)
    username_entry.pack(pady=5)

    create_label(forgot_window, "Secret Key")
    secret_key_frame = Frame(forgot_window)
    secret_key_frame.pack(pady=5)

    secret_key_entry = Entry(secret_key_frame, width=32, show="*")
    secret_key_entry.pack(side="left", padx=(0, 5))

    toggle_secret_key_button = Button(secret_key_frame, text="Show",
        command=lambda: toggle_visibility(secret_key_entry, toggle_secret_key_button),
        bg='#FF3030', fg='white', relief="flat", width=6)
    toggle_secret_key_button.pack(side="right")

    create_label(forgot_window, "New Password")
    new_password_frame = Frame(forgot_window)
    new_password_frame.pack(pady=5)

    new_password_entry = Entry(new_password_frame, width=32, show="*")
    new_password_entry.pack(side="left", padx=(0, 5))
    new_password_entry.bind("<KeyRelease>", update_strength)

    toggle_new_password_button = Button(new_password_frame, text="Show",
        command=lambda: toggle_visibility(new_password_entry, toggle_new_password_button),
        bg='#FF3030', fg='white', relief="flat", width=6)
    toggle_new_password_button.pack(side="right")

    # Password strength meter
    

    # Reset Button
    reset_button = Button(forgot_window, text="Reset Password", command=reset_password,
                          bg="#FF3030", fg="white", font=("Arial", 12), relief="flat")
    reset_button.pack(pady=10)
    strength_var = tk.DoubleVar()
    strength_bar = ttk.Progressbar(forgot_window, variable=strength_var, maximum=100, length=250)
    strength_bar.pack(pady=5)

    strength_label = tk.Label(forgot_window, text="", fg="white", bg="#2E2E2E")
    strength_label.pack()

def check_password(username, password, login_window):
    global current_user

    user_dir = get_user_path(username)
    user_data_file = os.path.join(user_dir, "data.enc")

    if os.path.exists(user_data_file):
        try:
            user_data = decrypt_data(user_data_file)

            if user_data is None:
                show_error_toast("❌ Error: Failed to decrypt user data. It may be corrupted.")
                return

            stored_username = user_data.get("username")
            stored_password_hash = user_data.get("password")
            stored_pin_hash = user_data.get("pin")

            if not stored_username:
                show_error_toast("❌ Error: Corrupted user data!")
                return

            if username != stored_username:
                show_error_toast("❌ Error: Invalid Username!")
                return

            if not stored_password_hash:
                show_error_toast("❌ Error: Password not set!")
                return

            if bcrypt.checkpw(password.encode(), stored_password_hash.encode()):
                current_user = username
                login_window.destroy()

                # If no PIN is set, prompt the user to create one
                if not stored_pin_hash:
                    show_warningtoast("⚠️ No PIN set for this account! Please create a PIN.")
                    prompt_set_pin(username)
                else:
                    show_success_toast(f"✅ Password Login Successful for {username}!")
                    main_app()
            else:
                show_error_toast("❌ Error: Incorrect Password!")

        except Exception as e:
            show_error_toast(f"❌ Error: {str(e)}")
    else:
        show_error_toast("❌ Error: User Not Found")


def check_pin(username, pin, login_window):
    global current_user

    user_dir = get_user_path(username)
    user_data_file = os.path.join(user_dir, "data.enc")

    if os.path.exists(user_data_file):
        try:
            user_data = decrypt_data(user_data_file)

            if user_data is None:
                show_error_toast("❌ Error: Failed to decrypt user data. It may be corrupted.")
                return

            stored_username = user_data.get("username")
            stored_pin_hash = user_data.get("pin")

            if not stored_username:
                show_error_toast("❌ Error: Corrupted user data!")
                return

            if username != stored_username:
                show_error_toast("❌ Error: Invalid Username!")
                return

            if not stored_pin_hash:
                show_warning_toast("⚠️ No PIN set for this account! Please login using Password.")
                return

            if bcrypt.checkpw(pin.encode(), stored_pin_hash.encode()):
                current_user = username
                login_window.destroy()
                show_success_toast(f"✅ PIN Login Successful for {username}!")
                main_app()
            else:
                show_error_toast("❌ Error: Incorrect PIN!")

        except Exception as e:
            show_error_toast(f"❌ Error: {str(e)}")
    else:
        show_error_toast("❌ Error: User Not Found")




def encrypt_text(text, key):
    cipher = Fernet(key)
    return cipher.encrypt(text.encode()).decode()

def decrypt_text(encrypted_text, key):
    cipher = Fernet(key)
    return cipher.decrypt(encrypted_text.encode()).decode()

# Function to save secret notes
def save_note(note_text, note_window):
    global current_user
    key = load_key()
    note = note_text.get("1.0", tk.END).strip()
    if not note:
        show_error_toast("❌ Error: Note can't be empty!")
        return

    encrypted_note = encrypt_text(note, key)
    username = current_user
    user_dir = get_user_path(username)
    notes_file_path = os.path.join(user_dir, "notes.enc")

    with open(notes_file_path, "a") as file:
        file.write(encrypted_note + "\n")

    show_success_toast("✅ Note Saved!")
    note_text.delete("1.0", tk.END)
    note_window.destroy()


# Function to view secret notes
def delete_note(note_listbox, decrypted_notes, encrypted_notes):
    global current_user
    selected_index = note_listbox.curselection()
    if not selected_index:
        show_error_toast("❌ Error: No note selected to delete!")
        return

    index = selected_index[0]

    # Remove from both decrypted and encrypted lists
    del decrypted_notes[index]
    del encrypted_notes[index]

    username = current_user
    user_dir = get_user_path(username)
    notes_file_path = os.path.join(user_dir, "notes.enc")

    # Save updated encrypted notes
    with open(notes_file_path, "w") as file:
        for line in encrypted_notes:
            file.write(line.strip() + "\n")

    note_listbox.delete(index)
    show_success_toast("✅ Note Deleted!")

# Function to view and delete secret notes
def view_notes():
    global current_user
    key = load_key()
    username = current_user
    user_dir = get_user_path(username)
    notes_file_path = os.path.join(user_dir, "notes.enc")

    if not os.path.exists(notes_file_path):
        show_error_toast("❌ Error: No saved notes found!")
        return

    with open(notes_file_path, "r") as file:
        encrypted_notes = [line.strip() for line in file if line.strip()]

    decrypted_notes = [decrypt_text(note, key) for note in encrypted_notes]

    notes_window = tk.Toplevel(root)
    notes_window.title("Saved Notes")
    notes_window.geometry("600x400")
    notes_window.configure(bg="black")
    notes_window.resizable(False, False)

    note_listbox = Listbox(notes_window, width=50, height=10, bg="black", fg="white")
    note_listbox.pack(pady=2, fill=tk.BOTH, expand=True)

    for note in decrypted_notes:
        note_listbox.insert(tk.END, note)

    delete_button = tk.Button(
        notes_window,
        text="Delete Selected Note",
        command=lambda: delete_note(note_listbox, decrypted_notes, encrypted_notes),
        bg="#CD853F",
        fg="white"
    )
    delete_button.pack(pady=2)

def export_notes():
    global current_user
    key = load_key()
    username = current_user
    user_dir = get_user_path(username)
    notes_file_path = os.path.join(user_dir, "notes.enc")

    if not os.path.exists(notes_file_path):
        show_error_toast("❌ Error: No saved notes found!")
        return

    with open(notes_file_path, "r") as file:
        encrypted_notes = [line.strip() for line in file if line.strip()]

    decrypted_notes = [decrypt_text(note, key) for note in encrypted_notes]

    # Prompt user to choose save location
    file_path = filedialog.asksaveasfilename(
        defaultextension=".txt",
        filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        title="Save Notes As"
    )

    if file_path:
        try:
            with open(file_path, "w", encoding="utf-8") as file:
                for note in decrypted_notes:
                    file.write(note + "\n")
            show_success_toast("✅ Notes exported Succcessfully ")
        except Exception as e:
            show_error_toast(f"❌ Error exporting notes: {e}")


# Function to create note window
def create_note():
    note_window = tk.Toplevel(root)
    note_window.title("Create Secret Note")
    note_window.geometry("600x400")
    note_window.configure(bg="black")
    note_window.attributes('-alpha', 0.9)
    note_window.resizable(False, False)
    
    tk.Label(note_window, text="Enter your secret note:", bg="black", fg="white").pack(pady=2)
    note_text = scrolledtext.ScrolledText(note_window, width=40, height=5, wrap=tk.WORD, bg="black", fg="white", insertbackground="white")
    note_text.pack(pady=2, fill=tk.BOTH, expand=True)
    
    tk.Button(note_window, text="Save Note", command=lambda: save_note(note_text, note_window), bg="#FF724C", fg="white").pack(pady=2)


def open_login_window():
    login_window = tk.Tk()
    login_window.title("Password Manager 5.1")
    login_window.geometry("600x400")
    login_window.configure(bg='#2E2E2E')
    login_window.resizable(False, False)

    login_window.grid_rowconfigure(0, weight=1)
    login_window.grid_columnconfigure(1, weight=1)

    # Left panel
    left_frame = tk.Frame(login_window, bg='#2E2E2E', width=180)
    left_frame.grid(row=0, column=0, sticky="ns", padx=15, pady=15)

    tk.Button(left_frame, text="Create User", command=register, bg='#BF3EFF', fg='white',
              font=("Arial", 12), relief='flat', activebackground='#005A9E', width=20, height=2).pack(pady=15, padx=10)

    tk.Button(left_frame, text="Forgot Password", command=forgot_password, bg='#FF3030', fg='black',
              font=("Arial", 12), relief='flat', activebackground='#E0A800', width=20, height=2).pack(pady=15, padx=10)

    tk.Button(left_frame, text="Generate Password", command=show_generated_password, bg='#0078D7', fg='black',
              font=("Arial", 12), relief='flat', activebackground='#E0A800', width=20, height=2).pack(pady=15, padx=10)

    tk.Button(left_frame, text="Admin Login", command=lambda: switch_to_admin_login(login_window),
              bg='#FFA500', fg='black', font=("Arial", 12), relief='flat',
              activebackground='#FF8C00', width=20, height=2).pack(pady=15, padx=10)

    tk.Label(left_frame, text="Created By:- Vikash Agnihotry", fg='white', bg='#2E2E2E',
             font=("Arial", 10, "italic")).pack(side="bottom", pady=2)

    # Separator
    separator = tk.Canvas(login_window, width=3, bg="black", highlightthickness=0)
    separator.grid(row=0, column=1, sticky="ns")

    # Right panel (Login Form)
    content_frame = tk.Frame(login_window, bg='#2E2E2E')
    content_frame.grid(row=0, column=2, sticky="nsew", padx=25, pady=20)

    tk.Label(content_frame, text="Existing User Login", fg='white', bg='#2E2E2E',
             font=("Arial", 14, "bold")).pack(pady=10)

    # Username entry
    tk.Label(content_frame, text="User Name", fg='white', bg='#2E2E2E', font=("Arial", 12)).pack(pady=5)
    username_entry = tk.Entry(content_frame, width=30)
    username_entry.pack(fill="x", padx=10, pady=5)

    # Login Mode Frame (for Password or PIN)
    login_mode_frame = tk.Frame(content_frame, bg='#2E2E2E')
    login_mode_frame.pack(pady=10)

    is_pin_mode = tk.BooleanVar(value=False)

    # Password section
    password_label = tk.Label(login_mode_frame, text="Password", fg='white', bg='#2E2E2E', font=("Arial", 12))
    password_frame = tk.Frame(login_mode_frame, bg='#2E2E2E')
    password_entry = tk.Entry(password_frame, width=25, show="*")

    def toggle_password_visibility():
        if password_entry.cget('show') == "*":
            password_entry.config(show="")
            toggle_pass_btn.config(text="Hide")
        else:
            password_entry.config(show="*")
            toggle_pass_btn.config(text="Show")

    toggle_pass_btn = tk.Button(password_frame, text="Show", command=toggle_password_visibility,
                                bg="#28A745", fg="white", relief="flat", width=5)

    # PIN section
    pin_label = tk.Label(login_mode_frame, text="PIN (4-6 digits)", fg='white', bg='#2E2E2E', font=("Arial", 12))
    pin_frame = tk.Frame(login_mode_frame, bg='#2E2E2E')
    pin_entry = tk.Entry(pin_frame, width=25, show="*")

    def toggle_pin_visibility():
        if pin_entry.cget('show') == "*":
            pin_entry.config(show="")
            toggle_pin_btn.config(text="Hide")
        else:
            pin_entry.config(show="*")
            toggle_pin_btn.config(text="Show")

    toggle_pin_btn = tk.Button(pin_frame, text="Show", command=toggle_pin_visibility,
                               bg="#28A745", fg="white", relief="flat", width=5)

    # Pack password login by default
    password_label.pack()
    password_frame.pack(pady=5)
    password_entry.pack(side="left", expand=True, fill="x", padx=(0, 5))
    toggle_pass_btn.pack(side="right")

    # Toggle between Password and PIN login
    def toggle_login_mode():
        if is_pin_mode.get():
            # Switch to password
            is_pin_mode.set(False)
            pin_label.pack_forget()
            pin_frame.pack_forget()
            password_label.pack()
            password_frame.pack(pady=5)
            toggle_mode_button.config(text="Switch to PIN Login")
        else:
            # Switch to PIN
            is_pin_mode.set(True)
            password_label.pack_forget()
            password_frame.pack_forget()
            pin_label.pack()
            pin_frame.pack(pady=5)
            pin_entry.pack(side="left", expand=True, fill="x", padx=(0, 5))
            toggle_pin_btn.pack(side="right")
            toggle_mode_button.config(text="Switch to Password Login")

    toggle_mode_button = tk.Button(content_frame, text="Switch to PIN Login",
                                   command=toggle_login_mode, bg="#0078D7", fg="white", font=("Arial", 11))
    toggle_mode_button.pack(pady=5)

    # Perform login
    def perform_login():
        username = username_entry.get().strip()
        if is_pin_mode.get():
            pin = pin_entry.get().strip()
            check_pin(username, pin, login_window)
        else:
            password = password_entry.get().strip()
            check_password(username, password, login_window)

    # Login Button
    tk.Button(content_frame, text="Login", command=perform_login,
              bg='#28A745', fg='white', font=("Arial", 12), relief='flat', activebackground='#218838', width=25).pack(pady=10)

    # Forgot PIN / Forgot Password options
   

    login_window.mainloop()

def search_password():
    user_dir = get_user_path(current_user)
    pass_file = os.path.join(user_dir, "pass.enc")
    data = decrypt_data(pass_file)
    
    search_window = Toplevel()
    search_window.title("Search Password")
    search_window.geometry("400x300")
    search_window.configure(bg='#2E2E2E')
    
    tk.Label(search_window, text="Enter Username", fg="white", bg="#2E2E2E", font=("Arial", 12)).pack(pady=5)
    search_entry = Entry(search_window, width=40)
    search_entry.pack(pady=5)
    
    def find_password():
        search_term = search_entry.get()
        results = []
        for website, credentials_list in data.items():
            if isinstance(credentials_list, list):
                for credentials in credentials_list:
                    if isinstance(credentials, dict) and credentials.get("username") == search_term:
                        decrypted_password = cipher_suite.decrypt(credentials["password"].encode()).decode()
                        results.append(f"Website: {website}\nPassword: {decrypted_password}\n{'-'*40}\n")
        
        result_text.delete(1.0, END)
        if results:
            result_text.insert(END, "\n".join(results))
        else:
            result_text.insert(END, "No passwords found for the given username.")

    create_button(search_window, "Search", find_password).pack(pady=5)
    result_text = Text(search_window, height=20, width=50, bg='#333', fg='white', font=("Arial", 12), relief='flat')
    result_text.pack(pady=5)



def delete_user():
    def confirm_deletion():
        username = username_entry.get().strip()
        secret_key = secret_key_entry.get().strip()
        
        if not username or not secret_key:
            show_error_toast("❌ Error: Fill All Fields!")
            return
        
        user_dir = get_user_path(username)
        user_data_file = os.path.join(user_dir, "data.enc")
        
        if not os.path.exists(user_data_file):
            show_error_toast("❌ Error: User Not Found!")
            return

        try:
            user_data = decrypt_data(user_data_file)
            stored_secret_key = cipher_suite.decrypt(user_data["secret_key"].encode()).decode()
        except (InvalidToken, KeyError, AttributeError):
            show_error_toast("❌ Error: Invalid/Currupted Data!")
            return
        
        if secret_key != stored_secret_key:
            show_error_toast("❌ Error: Incorrect Secret Key!")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete your account? Confirm all passwords are Exported & Data Decrypted."):
            try:
                shutil.rmtree(user_dir)
                messagebox.showinfo("Deleted", "User account and all associated data have been deleted successfully.")
                delete_window.destroy()
                logout()
            except Exception as e:
                show_error_toast(f"Failed to delete user: {str(e)}")
    
    delete_window = Toplevel()
    delete_window.title("Delete User Account")
    delete_window.geometry("400x350")
    delete_window.configure(bg='#2E2E2E')

    Label(delete_window, text="Username", fg='white', bg='#2E2E2E', font=("Arial", 12)).pack(pady=5)
    username_entry = Entry(delete_window, width=40)
    username_entry.pack(pady=5)
    username_entry.focus()

    Label(delete_window, text="Secret Key", fg='white', bg='#2E2E2E', font=("Arial", 12)).pack(pady=5)
    
    # Secret Key Entry with Show/Hide button
    secret_key_frame = Frame(delete_window, bg='#2E2E2E')
    secret_key_frame.pack(pady=5)

    secret_key_entry = Entry(secret_key_frame, width=32, show="*")
    secret_key_entry.pack(side="left", padx=(0, 5))

    def toggle_secret_key():
        if secret_key_entry.cget("show") == "*":
            secret_key_entry.config(show="")
            toggle_secret_key_button.config(text="Hide")
        else:
            secret_key_entry.config(show="*")
            toggle_secret_key_button.config(text="Show")

    toggle_secret_key_button = Button(secret_key_frame, text="Show", command=toggle_secret_key, 
                                      bg='#D9534F', fg='white', relief="flat", width=6)
    toggle_secret_key_button.pack(side="right")

    # Buttons for Deletion and Cancel
    Button(delete_window, text="Confirm Deletion", command=confirm_deletion, 
           bg='#D9534F', fg='white', font=("Arial", 12), relief='flat', activebackground='#C9302C').pack(pady=10)

    Button(delete_window, text="Cancel", command=delete_window.destroy, 
           bg='#28A745', fg='white', font=("Arial", 12), relief='flat').pack(pady=5)

def import_from_excel():
    user_dir = get_user_path(current_user)
    pass_file = os.path.join(user_dir, "pass.enc")

    # Open file dialog to select Excel file
    file_path = filedialog.askopenfilename(
        title="Select Password Excel File",
        filetypes=[("Excel files", "*.xlsx;*.xls")]
    )

    if not file_path:
        return  # User canceled file selection

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        data = decrypt_data(pass_file) or {}  # Ensure we have a dictionary

        imported_count, updated_count, skipped_count = 0, 0, 0
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is headers
            website, username, password = row[:3]  # First three columns

            if not (website and username and password):
                continue  # Skip invalid rows

            encrypted_password = cipher_suite.encrypt(password.encode()).decode()

            # Check if exact same entry exists (skip)
            for existing_website, entries in data.items():
                if isinstance(entries, dict):  # Convert old format to list
                    data[existing_website] = [entries]

                if existing_website == website:
                    for entry in data[existing_website]:
                        if entry["username"] == username:
                            if entry["password"] == encrypted_password:
                                skipped_count += 1  # No change, skip
                                break
                            else:
                                # Password changed, ask user if they want to update
                                update = messagebox.askyesno(
                                    "Duplicate Entry",
                                    f"Username '{username}' for '{website}' exists with a different password. Update?"
                                )
                                if update:
                                    entry["password"] = encrypted_password
                                    updated_count += 1
                                else:
                                    skipped_count += 1
                                break
                    else:
                        # New username, create a separate website entry
                        unique_website = website
                        count = 1
                        while unique_website in data:
                            unique_website = f"{website} ({count})"
                            count += 1

                        data[unique_website] = [{"username": username, "password": encrypted_password}]
                        imported_count += 1
                    break
            else:
                # Completely new website entry
                data[website] = [{"username": username, "password": encrypted_password}]
                imported_count += 1

        encrypt_data(data, pass_file)  # Save updated password data

        messagebox.showinfo(
            "Import Complete",
            f"Passwords Imported: {imported_count}\nUpdated: {updated_count}\nSkipped: {skipped_count}"
        )

    except Exception as e:
        show_error_toast( f"An error occurred: {str(e)}")

def export_passwords_to_excel():
    try:
        if not current_user:
            show_error_toast("Error: No user logged in.")
            return

        user_dir = get_user_path(current_user)
        pass_file = os.path.join(user_dir, "pass.enc")  # User-specific password file

        if not os.path.exists(pass_file):
            show_error_toast("Error: No stored passwords found.")
            return

        # Decrypt stored passwords
        password_data = decrypt_data(pass_file)

        # Ask user where to save the file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Passwords as Excel"
        )

        if not file_path:
            return  # User canceled file selection

        # Create and save Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"{current_user} Passwords"

        # Save passwords
        sheet.append(["Website", "Username", "Password"])
        for website, credentials_list in password_data.items():
            for credentials in credentials_list:
                decrypted_password = cipher_suite.decrypt(credentials["password"].encode()).decode()
                sheet.append([website, credentials["username"], decrypted_password])

        workbook.save(file_path)
        show_success_toast("✅Password Successfully Exported!")

    except Exception as e:
        messagebox.showerror("Export Error", f"An error occurred: {str(e)}")
def get_encrypted_files_path():
    """Returns the path for the user's encrypted files directory."""
    encrypted_dir = os.path.join(get_user_path(current_user), "Encrypted Files")
    os.makedirs(encrypted_dir, exist_ok=True)
    return encrypted_dir

def get_decrypted_files_path():
    """Returns the path for the user's decrypted files directory."""
    decrypted_dir = os.path.join(get_user_path(current_user), "Decrypted Files")
    os.makedirs(decrypted_dir, exist_ok=True)
    return decrypted_dir

def get_encrypted_list_path():
    """Returns the path for the encrypted file list (list.enc)."""
    return os.path.join(get_user_path(current_user), "list.enc")

def save_encrypted_list(file_list):
    """Encrypt and save the encrypted file list."""
    encrypt_data(file_list, get_encrypted_list_path())

def load_encrypted_list():
    """Load and decrypt the encrypted file list."""
    return decrypt_data(get_encrypted_list_path()) if os.path.exists(get_encrypted_list_path()) else []

import os
from tkinter import messagebox, filedialog

def encrypt_files():
    """Encrypt selected files and save them in the user's 'Encrypted Files' folder."""
    try:
        if not current_user:
            show_error_toast("No user logged in.")
            return

        file_paths = filedialog.askopenfilenames(title="Select files to encrypt")
        if not file_paths:
            return

        encrypted_dir = get_encrypted_files_path()
        if not os.path.exists(encrypted_dir):
            os.makedirs(encrypted_dir)  # Create directory if missing

        if 'cipher_suite' not in globals():
             show_error_toast("❌ Error: Encryption System Not Initialized ")
             return

        encrypted_files = load_encrypted_list()  # Load existing encrypted file list

        for file_path in file_paths:
            try:
                encrypted_file_name = os.path.basename(file_path) + ".encr"
                encrypted_file_path = os.path.join(encrypted_dir, encrypted_file_name)

                with open(file_path, "rb") as f:
                    encrypted_data = cipher_suite.encrypt(f.read())

                with open(encrypted_file_path, "wb") as f:
                    f.write(encrypted_data)

                if encrypted_file_name not in encrypted_files:
                    encrypted_files.append(encrypted_file_name)

                # Optional: Remove original file after encryption
                os.remove(file_path)

            except Exception as e:
                show_warning_toast(f"Failed to encrypt {os.path.basename(file_path)}: {str(e)}")

        save_encrypted_list(encrypted_files)  # Save updated list

        show_success_toast(f"{len(file_paths)} files encrypted successfully!")

    except Exception as e:
        show_error_toast("Error", f"An error occurred: {str(e)}")


def decrypt_file():
    """Show a list of encrypted files for the user to select and decrypt."""
    try:
        if not current_user:
            show_error_toast("Error", "No user logged in.")
            return

        encrypted_files = load_encrypted_list()
        if not encrypted_files:
            show_warning_toast("⚠No Encrypted File Found!")
            return

        def on_select():
            """Decrypt the selected file."""
            selected_index = listbox.curselection()
            if not selected_index:
               show_warning_toast("Selection Required", "Please select a file to decrypt.")
               return

            file_to_decrypt = encrypted_files[selected_index[0]]
            encrypted_dir = get_encrypted_files_path()
            decrypted_dir = get_decrypted_files_path()

            encrypted_file_path = os.path.join(encrypted_dir, file_to_decrypt)
            decrypted_file_path = os.path.join(decrypted_dir, file_to_decrypt.replace(".encr", ""))

            with open(encrypted_file_path, "rb") as f:
                decrypted_data = cipher_suite.decrypt(f.read())

            with open(decrypted_file_path, "wb") as f:
                f.write(decrypted_data)

            # Remove file from encrypted list
            encrypted_files.remove(file_to_decrypt)
            save_encrypted_list(encrypted_files)
            os.remove(encrypted_file_path)

            show_success_toast(f"File '{file_to_decrypt}' decrypted successfully!")
            decrypt_window.destroy()  # Close the selection window

        # Create selection window
        decrypt_window = Toplevel()
        decrypt_window.title("Select File to Decrypt")
        decrypt_window.geometry("400x300")
        decrypt_window.configure(bg="#2E2E2E")

        listbox = Listbox(decrypt_window, bg="#333", fg="white", font=("Arial", 12), relief="flat")
        listbox.pack(fill="both", expand=True, padx=10, pady=10)

        for file in encrypted_files:
            listbox.insert(END, file)

        Button(decrypt_window, text="Decrypt Selected", command=on_select, bg='#48D1CC', fg='white',
               font=("Arial", 12), relief="flat", activebackground='#C9302C').pack(pady=10)

    except Exception as e:
        show_error_toast(f"An error occurred: {str(e)}")
        
def view_encrypted_files():
    """Display all encrypted files stored in list.enc."""
    try:
        encrypted_files = load_encrypted_list()
        if not encrypted_files:
            show_warning_toast("⚠No Encrypted Files Found!")
            return

        top = Toplevel()
        top.title("Encrypted Files")
        top.geometry("400x300")
        top.configure(bg="#2E2E2E")

        listbox = Listbox(top, bg="#333", fg="white", font=("Arial", 12), relief="flat")
        listbox.pack(fill="both", expand=True)

        for file in encrypted_files:
            listbox.insert(END, file)

    except Exception as e:
        show_error_toast(f"An error occurred: {str(e)}")

def decrypt_all_files():
    """Decrypt all files from 'Encrypted Files' to 'Decrypted Files'."""
    try:
        if not current_user:
            show_error_toast("❌ No user logged in!")
            return

        encrypted_dir = get_encrypted_files_path()
        decrypted_dir = get_decrypted_files_path()

        encrypted_files = load_encrypted_list()
        if not encrypted_files:
            show_warning_toast("⚠️ No encrypted files found.")
            return

        decrypted_count = 0

        for file_name in encrypted_files.copy():
            encrypted_file_path = os.path.join(encrypted_dir, file_name)
            decrypted_file_path = os.path.join(decrypted_dir, file_name.replace(".encr", ""))

            if os.path.exists(encrypted_file_path):
                try:
                    with open(encrypted_file_path, "rb") as f:
                        decrypted_data = cipher_suite.decrypt(f.read())
                    with open(decrypted_file_path, "wb") as f:
                        f.write(decrypted_data)

                    os.remove(encrypted_file_path)
                    encrypted_files.remove(file_name)
                    decrypted_count += 1

                except Exception as e:
                    show_error_toast(f"❌ Failed to decrypt {file_name}: {e}")

        save_encrypted_list(encrypted_files)  # Update encrypted file list after decryption

        if decrypted_count > 0:
            show_success_toast(f"✅ {decrypted_count} files decrypted successfully!")
        else:
            show_warning_toast("⚠️ No files decrypted.")

    except Exception as e:
        show_error_toast(f"❌ Error during decryption: {e}")

       
def on_exit():
    """Ask the user for confirmation before exiting"""
    if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
        root.destroy()  # Close the application      

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Toplevel
from collections import defaultdict
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import re
import openpyxl

def export_to_pdf():
    file_path = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        title="Export Password Report to PDF"
    )
    if not file_path:
        return

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Password Health Report", ln=True, align="C")
    pdf.ln(10)

    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Passwords Summary:", ln=True)
    pdf.set_font("Arial", size=11)
    for item in tree.get_children():
        website, username, strength, status = tree.item(item, "values")
        pdf.multi_cell(0, 10, f"Website: {website}\nUsername: {username}\nStrength: {strength}\nStatus: {status}\n", border=0)
    try:
        pdf.output(file_path)
        show_success_toast("✅PDF Report Saved Successfully !")
    except Exception as e:
        show_error_toast(f"Failed to export PDF: {str(e)}")

def show_password_health_dashboard():
    global dashboard, tree, entry_details  # make accessible to other functions
    user_dir = get_user_path(current_user)
    pass_file = os.path.join(user_dir, "pass.enc")
    if not os.path.exists(pass_file):
         show_error_toast("❌ Error: Data File Not Found ")
         return

    data = decrypt_data(pass_file)
    if not data:
         show_error_toast("❌ No Saved Password Found ")
         return

    dashboard = Toplevel()
    dashboard.title("Password Health Dashboard")
    dashboard.geometry("900x850")
    dashboard.configure(bg="#1E1E1E")

    style = ttk.Style(dashboard)
    style.theme_use("clam")
    style.configure("Treeview", background="#2E2E2E", foreground="white", fieldbackground="#2E2E2E", font=("Arial", 10), anchor="center")
    style.configure("Treeview.Heading", font=("Arial", 11, "bold"), foreground="white", background="#333", anchor="center")

    tree = ttk.Treeview(dashboard, columns=("Website", "Username", "Strength", "Status"), show="headings")
    for col in ("Website", "Username", "Strength", "Status"):
        tree.heading(col, text=col, anchor="center")
        tree.column(col, anchor="center")
    tree.pack(fill="both", expand=True, padx=10, pady=10)

    tree.tag_configure("weak", background="#8B0000")
    tree.tag_configure("reused", background="#B8860B")
    tree.tag_configure("strong", background="#006400")

    password_map = defaultdict(list)
    all_entries = []
    strength_counts = {"Weak": 0, "Moderate": 0, "Strong": 0}
    weak = reused = 0
    entry_details = {}

    def evaluate_strength(pwd):
        score = 0
        if len(pwd) >= 8:
            score += 1
        if re.search(r"[a-z]", pwd) and re.search(r"[A-Z]", pwd):
            score += 1
        if re.search(r"\d", pwd):
            score += 1
        if re.search(r"[!@#$%^&*(),.?\":{}|<>]", pwd):
            score += 1
        return "Weak" if score <= 1 else "Moderate" if score <= 3 else "Strong"

    for website, creds_list in data.items():
        for creds in creds_list:
            decrypted_password = cipher_suite.decrypt(creds["password"].encode()).decode()
            password_map[decrypted_password].append((website, creds["username"]))
            all_entries.append((website, creds["username"], decrypted_password))

    for website, username, password in all_entries:
        strength = evaluate_strength(password)
        strength_counts[strength] += 1

        is_reused = len(password_map[password]) > 1
        status = []
        if strength == "Weak":
            weak += 1
            status.append("⚠️ Weak")
        if is_reused:
            reused += 1
            status.append("🔁 Reused")

        final_status = ", ".join(status) if status else "✓ Good"
        tag = "weak" if strength == "Weak" else "reused" if is_reused else "strong"

        item_id = tree.insert("", "end", values=(website, username, strength, final_status), tags=(tag,))
        entry_details[item_id] = password

    summary = tk.Label(dashboard, text=f"Total Passwords: {len(all_entries)}   Weak: {weak}   Reused: {reused}",
                       font=("Arial", 13, "bold"), bg="#1E1E1E", fg="lightgray")
    summary.pack(pady=10)

    def show_entry_details():
        selected_item = tree.focus()
        if not selected_item:
            return

        website, username, strength, status = tree.item(selected_item, "values")
        password = entry_details.get(selected_item, "Unknown")

        details_window = Toplevel(dashboard)
        details_window.title("Password Entry Details")
        details_window.geometry("400x500")
        details_window.configure(bg="#2E2E2E")

        tk.Label(details_window, text=f"Website: {website}", font=("Arial", 12), bg="#2E2E2E", fg="white").pack(pady=5)
        tk.Label(details_window, text=f"Username: {username}", font=("Arial", 12), bg="#2E2E2E", fg="white").pack(pady=5)

        password_label = tk.Label(details_window, text="Password: 🔒 ******", font=("Arial", 12), bg="#2E2E2E", fg="white")
        password_label.pack(pady=5)

        def toggle_password():
            if show_password_var.get():
                password_label.config(text=f"Password: {password}")
            else:
                password_label.config(text="Password: 🔒 ******")

        show_password_var = tk.BooleanVar(value=False)
        tk.Checkbutton(details_window, text="Show Password", variable=show_password_var, command=toggle_password,
                       font=("Arial", 11), bg="#2E2E2E", fg="white", selectcolor="#2E2E2E").pack(pady=5)

        tk.Label(details_window, text=f"Strength: {strength}", font=("Arial", 12), bg="#2E2E2E", fg="white").pack(pady=5)
        tk.Label(details_window, text=f"Status: {status}", font=("Arial", 12), bg="#2E2E2E", fg="white").pack(pady=5)

        def copy_password():
            dashboard.clipboard_clear()
            dashboard.clipboard_append(password)
            dashboard.update()
            show_success_toast("✅Password Copied!")

        def copy_username():
            dashboard.clipboard_clear()
            dashboard.clipboard_append(username)
            dashboard.update()
            show_success_toast("✅UserName Copied!")

        btn_frame = tk.Frame(details_window, bg="#2E2E2E")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="📋 Copy Username", command=copy_username, bg="#0078D7", fg="white", font=("Arial", 11, "bold"), relief="flat").pack(side="left", padx=10)
        tk.Button(btn_frame, text="📋 Copy Password", command=copy_password, bg="#0078D7", fg="white", font=("Arial", 11, "bold"), relief="flat").pack(side="left", padx=10)

        tk.Button(details_window, text="Close", command=details_window.destroy, bg="#555", fg="white", font=("Arial", 11), relief="flat").pack(pady=15)

    def delete_selected_entry():
        selected_item = tree.focus()
        if not selected_item:
            return
        confirm = messagebox.askyesno("Delete", "Are you sure you want to delete this entry?")
        if confirm:
            tree.delete(selected_item)
            if selected_item in entry_details:
                del entry_details[selected_item]

    # Setup context menu
    context_menu = tk.Menu(tree, tearoff=0, bg="#333", fg="white", font=("Arial", 10))
    context_menu.add_command(label="🔍 View Details", command=show_entry_details)
    context_menu.add_separator()
    context_menu.add_command(label="🗑 Delete Entry", command=delete_selected_entry)

    def on_right_click(event):
        region = tree.identify("region", event.x, event.y)
        if region == "cell":
            iid = tree.identify_row(event.y)
            if iid:
                tree.selection_set(iid)
                tree.focus(iid)
                context_menu.tk_popup(event.x_root, event.y_root)

    tree.bind("<Button-3>", on_right_click)

    def export_dashboard():
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Password Health Report"
        )
        if not file_path:
            return
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Password Health Report"
            sheet.append(["Website", "Username", "Strength", "Status"])
            for row_id in tree.get_children():
                row = tree.item(row_id)["values"]
                sheet.append(row)
            workbook.save(file_path)
            show_success_toast("✅Report Saved successfully!")
        except Exception as e:
            show_error_toast(f"Export failed: {str(e)}")

    export_frame = tk.Frame(dashboard, bg="#1E1E1E")
    export_frame.pack(pady=10)

    export_report_button = tk.Button(export_frame, text="📄 Export to Excel", command=export_dashboard,
                                 bg="#0078D7", fg="white", font=("Arial", 11, "bold"), relief="flat")
    export_report_button.pack(side="left", padx=10)

    pdf_button = tk.Button(export_frame, text="🖨 Export to PDF", command=export_to_pdf,
                       bg="#0078D7", fg="white", font=("Arial", 11, "bold"), relief="flat")
    pdf_button.pack(side="left", padx=10)


    # Pie chart
    fig, ax = plt.subplots(figsize=(5.5, 5.5), dpi=100)
    labels = list(strength_counts.keys())
    sizes = list(strength_counts.values())
    colors = ["#8B0000", "#FF8C00", "#006400"]

    wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct="%1.1f%%", startangle=90, colors=colors)
    ax.axis("equal")
    ax.set_title("Password Strength Distribution", color="white", pad=40)

    def on_hover(event):
        for i, wedge in enumerate(wedges):
            if wedge.contains_point([event.x, event.y]):
                ax.set_title(f"{labels[i]}: {sizes[i]} password(s)", color="white", pad=40)
                fig.canvas.draw_idle()
                return
        ax.set_title("Password Strength Distribution", color="white", pad=40)
        fig.canvas.draw_idle()

    fig.canvas.mpl_connect("motion_notify_event", on_hover)

    pie_canvas = FigureCanvasTkAgg(fig, master=dashboard)
    pie_canvas.draw()
    pie_canvas.get_tk_widget().pack(pady=30)
    fig.patch.set_facecolor("#1E1E1E")


def decrypt_and_view_image():
    try:
        if not current_user:
            show_error_toast("No user logged in.")
            return

        file_path = filedialog.askopenfilename(
            title="Select Encrypted Media File",
            filetypes=[("Encrypted Files", "*.encr")]
        )
        if not file_path:
            return

        encrypted_dir = os.path.dirname(file_path)
        temp_dir = os.path.join(encrypted_dir, "temp")
        os.makedirs(temp_dir, exist_ok=True)

        original_name = os.path.basename(file_path)[:-5]  # Remove '.encr'
        temp_decrypted_path = os.path.join(temp_dir, original_name)

        with open(file_path, "rb") as f:
            decrypted_data = cipher_suite.decrypt(f.read())

        with open(temp_decrypted_path, "wb") as f:
            f.write(decrypted_data)

        if original_name.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff")):
            view_decrypted_image(temp_decrypted_path)
        elif original_name.lower().endswith((".mp4", ".avi", ".mov", ".mkv")):
            view_decrypted_video(temp_decrypted_path)
        else:
            messagebox.showerror("Unsupported File", "Only images or video files are supported.")

    except Exception as e:
        show_error_toast( f"An error occurred: {str(e)}")

    except Exception as e:
        show_error_toast(f"An error occurred: {str(e)}")

# Helper to open decrypted image

def view_decrypted_video(video_path):
    try:
        video_window = tk.Toplevel()
        video_window.title("Decrypted Video Viewer")
        video_window.geometry("400x200")
        video_window.configure(bg="#1E1E1E")

        tk.Label(video_window, text=os.path.basename(video_path), bg="#1E1E1E", fg="white", font=("Arial", 12)).pack(pady=10)

        def open_video():
            import webbrowser
            webbrowser.open(video_path)

        def save_decrypted_copy():
            save_path = filedialog.asksaveasfilename(
                defaultextension=".mp4",
                filetypes=[("Video Files", "*.mp4;*.avi;*.mov;*.mkv"), ("All Files", "*.*")]
            )
            if save_path:
                with open(video_path, "rb") as temp_file, open(save_path, "wb") as out_file:
                    out_file.write(temp_file.read())
                show_success_toast("Saved", "Decrypted video saved successfully!")

        btn_frame = tk.Frame(video_window, bg="#1E1E1E")
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="▶️ Play Video", command=open_video,
                  bg="#28a745", fg="white", font=("Arial", 11), relief="flat").pack(side="left", padx=10)
        tk.Button(btn_frame, text="💾 Download Decrypted", command=save_decrypted_copy,
                  bg="#0078D7", fg="white", font=("Arial", 11), relief="flat").pack(side="left", padx=10)

        def close_video_window():
            try:
                if os.path.exists(video_path):
                    os.remove(video_path)
            except Exception as e:
                print(f"Failed to delete temp video: {str(e)}")
            video_window.destroy()

        tk.Button(video_window, text="Close", command=close_video_window,
              bg="#555", fg="white", font=("Arial", 11), relief="flat").pack(pady=10)


    except Exception as e:
        show_error_toast(f"Failed to open decrypted video: {str(e)}")



def view_decrypted_image(image_path):
    try:
        img_window = tk.Toplevel()
        img_window.title("Decrypted Image Viewer")
        img_window.geometry("720x800")
        img_window.configure(bg="#1E1E1E")

        img = Image.open(image_path)

        canvas_frame = tk.Frame(img_window, bg="#1E1E1E")
        canvas_frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(canvas_frame, bg="#1E1E1E", highlightthickness=0)
        scroll_y = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scroll_x = tk.Scrollbar(canvas_frame, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)

        image_container = tk.Frame(canvas, bg="#1E1E1E")
        canvas.create_window((0, 0), window=image_container, anchor="nw")

        def update_scrollregion(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        image_container.bind("<Configure>", update_scrollregion)

        img_tk = ImageTk.PhotoImage(img)
        img_label = tk.Label(image_container, image=img_tk, bg="#1E1E1E")
        img_label.image = img_tk
        img_label.pack(pady=20)

        def save_decrypted_copy():
            save_path = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG Image", "*.png"), ("All Files", "*.*")]
            )
            if save_path:
                with open(image_path, "rb") as temp_file, open(save_path, "wb") as out_file:
                    out_file.write(temp_file.read())
                show_success_toast("Saved", "Decrypted image saved successfully!")

        btn_frame = tk.Frame(img_window, bg="#1E1E1E")
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="💾 Download Decrypted", command=save_decrypted_copy,
                  bg="#28a745", fg="white", font=("Arial", 11), relief="flat").pack(side="left", padx=10)
        def close_image_window():
            try:
                if os.path.exists(image_path):
                    os.remove(image_path)
            except Exception as e:
                print(f"Failed to delete temp image: {str(e)}")
            img_window.destroy()

        tk.Button(btn_frame, text="Close", command=close_image_window,
              bg="#555", fg="white", font=("Arial", 11), relief="flat").pack(side="left", padx=10)


    except Exception as e:
        show_error_toast(f"Failed to open decrypted image: {str(e)}")


# ========== Banking Details Manager ==========

def show_toast(message):
    def run_toast():
        toast = tk.Toplevel()
        toast.overrideredirect(True)
        toast.configure(bg="#333")

        toast.update_idletasks()
        width, height = 300, 40
        x = toast.winfo_screenwidth() - width - 20
        y = toast.winfo_screenheight() - height - 60
        toast.geometry(f"{width}x{height}+{x}+{y}")

        label = tk.Label(toast, text=message, bg="#333", fg="white", font=("Arial", 10))
        label.pack(fill="both", expand=True)

        toast.after(2000, toast.destroy)
        toast.mainloop()

    threading.Thread(target=run_toast).start()

# ========= Banking Details Manager =========

def show_toast(message):
    toast = tk.Toplevel()
    toast.overrideredirect(True)
    toast.configure(bg="#333")

    toast.update_idletasks()
    width, height = 300, 40
    x = toast.winfo_screenwidth() - width - 20
    y = toast.winfo_screenheight() - height - 60
    toast.geometry(f"{width}x{height}+{x}+{y}")

    label = tk.Label(toast, text=message, bg="#333", fg="white", font=("Arial", 10))
    label.pack(fill="both", expand=True)

    toast.after(2000, toast.destroy)

# ========= Banking Details Manager =========

def launch_banking_ui():
    banking_window = tk.Toplevel()
    banking_window.title("Banking Details")
    banking_window.geometry("500x650")
    banking_window.configure(bg="#1E1E1E")

    def toggle_fields():
        if is_credit_card.get():
            entry_account.config(state="disabled")
            entry_ifsc.config(state="disabled")
        else:
            entry_account.config(state="normal")
            entry_ifsc.config(state="normal")

    def load_banking_data():
        user_dir = get_user_path(current_user)
        file_path = os.path.join(user_dir, "bdata.enc")
        if not os.path.exists(file_path):
            return []
        try:
            with open(file_path, "rb") as f:
                decrypted = cipher_suite.decrypt(f.read()).decode()
            data = eval(decrypted)
            return [data] if isinstance(data, dict) else data
        except:
            return []

    def save_banking_data():
        bank_name = entry_bank.get()
        account_number = entry_account.get()
        ifsc_code = entry_ifsc.get()
        card_number = entry_card.get()
        card_expiry = entry_expiry.get()
        cvv = entry_cvv.get()
        account_type = account_type_var.get().strip()

        if not bank_name or not account_type:
            show_warning_toast("⚠️ Bank Name and Account Type are required!")
            return

        if not (bank_name and card_number and card_expiry):
            show_warning_toast("⚠️ Fill Mandatory Fields")
            return

        if not is_credit_card.get():
            if not (account_number and ifsc_code):
                show_warning_toast("⚠️ Account Number and IFSC are required.")
                return
            if not account_number.isdigit():
                show_warning_toast("⚠️ Account Number Must Be Numeric.")
                return
            if not (len(ifsc_code) == 11 and ifsc_code[:4].isalpha() and ifsc_code[4:].isdigit()):
                show_warning_toast("⚠️ IFSC Code Must be 11 Characters (4 Alpha + 7 Numeric).")
                return

        if not (card_number.isdigit() and len(card_number) == 16):
            show_warning_toast("⚠️ Card Number Must be 16 Digits.")
            return

        if not re.match(r"^(0[1-9]|1[0-2])/([0-9]{2})$", card_expiry):
            show_warning_toast("⚠️ Card Expiry must be in MM/YY format.")
            return

        if cvv and (not cvv.isdigit() or len(cvv) != 3):
            show_warning_toast("⚠️ CVV must be 3 digits if provided.")
            return

        new_entry = {
            "Account Type": account_type,
            "Bank Name": bank_name,
            "Account Number": account_number if not is_credit_card.get() else "",
            "IFSC Code": ifsc_code if not is_credit_card.get() else "",
            "Card Number": card_number,
            "Card Expiry": card_expiry,
            "CVV": cvv
        }

        user_dir = get_user_path(current_user)
        file_path = os.path.join(user_dir, "bdata.enc")
        banking_data = []

        if os.path.exists(file_path):
            try:
                with open(file_path, "rb") as f:
                    decrypted = cipher_suite.decrypt(f.read()).decode()
                banking_data = eval(decrypted)
                if isinstance(banking_data, dict):
                    banking_data = [banking_data]
            except Exception as e:
                show_error_toast(f"Failed to load data: {str(e)}")
                return

        for entry in banking_data:
            if entry.get("Card Number") == card_number:
                show_error_toast("❌ Error: Duplicate Card Number.")
                return

        banking_data.append(new_entry)

        try:
            encrypted = cipher_suite.encrypt(str(banking_data).encode())
            with open(file_path, "wb") as f:
                f.write(encrypted)
            show_success_toast("✅ Banking details saved successfully!")
            for widget in [entry_bank, entry_account, entry_ifsc, entry_card, entry_expiry, entry_cvv]:
                widget.delete(0, tk.END)
            refresh_view()
        except Exception as e:
            show_error_toast(f"Failed to save banking data: {str(e)}")

    def refresh_view():
        for widget in view_frame.winfo_children():
            widget.destroy()

        banking_data = load_banking_data()
        if not banking_data:
            tk.Label(view_frame, text="No saved banking details.", bg="#1E1E1E", fg="white", font=("Arial", 11)).pack(pady=20)
            return

        canvas = tk.Canvas(view_frame, bg="#1E1E1E", highlightthickness=0)
        scrollbar_y = tk.Scrollbar(view_frame, orient="vertical", command=canvas.yview)
        scrollbar_x = tk.Scrollbar(view_frame, orient="horizontal", command=canvas.xview)
        scrollable_frame = tk.Frame(canvas, bg="#1E1E1E")

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        canvas.pack(side="top", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        cards_row = tk.Frame(scrollable_frame, bg="#1E1E1E")
        cards_row.pack(padx=10, pady=10, fill="x")

        for idx, entry in enumerate(banking_data):
            card = tk.Frame(cards_row, bg="#333", bd=1, relief="ridge")
            card.pack(side="left", padx=10, pady=5)

            for key, value in entry.items():
                frame = tk.Frame(card, bg="#333")
                frame.pack(anchor="w", padx=5, pady=2, fill="x")
                tk.Label(frame, text=f"{key}: {value}", bg="#333", fg="white", font=("Arial", 10)).pack(side="left")
                if key in ("Account Number", "IFSC Code", "Card Number"):
                    def copy_to_clipboard(v=value):
                        banking_window.clipboard_clear()
                        banking_window.clipboard_append(v)
                        banking_window.update()
                        show_toast("Copied")
                    tk.Button(frame, text="📋", command=copy_to_clipboard, bg="#555", fg="white", font=("Arial", 9), relief="flat", width=3).pack(side="left", padx=5)

            btn_frame = tk.Frame(card, bg="#333")
            btn_frame.pack(pady=5)

            tk.Button(btn_frame, text="✏️ Edit", command=lambda i=idx: edit_entry(i), bg="#0078D7", fg="white", font=("Arial", 10), relief="flat").pack(side="left", padx=5)
            tk.Button(btn_frame, text="🗑 Delete", command=lambda i=idx: delete_entry(i), bg="#dc3545", fg="white", font=("Arial", 10), relief="flat").pack(side="left", padx=5)

    def export_to_excel():
        import openpyxl
        banking_data = load_banking_data()
        if not banking_data:
           show_warning_toast("No Data", "No banking details to export.")
           return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Banking Data"

        sheet.append(["Account Type", "Bank Name", "Account Number", "IFSC Code", "Card Number", "Card Expiry", "CVV"])

        for entry in banking_data:
            sheet.append([
                entry.get("Account Type", "Saving Account"),
                entry.get("Bank Name", ""),
                entry.get("Account Number", ""),
                entry.get("IFSC Code", ""),
                entry.get("Card Number", ""),
                entry.get("Card Expiry", ""),
                entry.get("CVV", "")
            ])

        workbook.save(file_path)
        show_success_toast("✅ Exported to Excel successfully!")

    def export_to_pdf():
        banking_data = load_banking_data()
        if not banking_data:
           show_warning_toast("No Data", "No banking details to export.")
           return

        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
        if not file_path:
            return

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Banking Details", ln=True, align="C")
        pdf.ln(10)

        for entry in banking_data:
            for key in ["Account Type", "Bank Name", "Account Number", "IFSC Code", "Card Number", "Card Expiry", "CVV"]:
                pdf.cell(0, 10, f"{key}: {entry.get(key, '')}", ln=True)
            pdf.ln(8)

        pdf.output(file_path)
        show_success_toast("✅ Exported to PDF successfully!")

    # --- UI Layout ---
    notebook = ttk.Notebook(banking_window)
    notebook.pack(fill="both", expand=True, pady=10)

    add_frame = tk.Frame(notebook, bg="#1E1E1E")
    view_frame = tk.Frame(notebook, bg="#1E1E1E")
    notebook.add(add_frame, text="➕ Add Details")
    notebook.add(view_frame, text="📄 View Details")

    is_credit_card = tk.BooleanVar()

    tk.Checkbutton(add_frame, text="Add Credit Card", variable=is_credit_card, bg="#1E1E1E", fg="white",
                   activebackground="#1E1E1E", activeforeground="white", selectcolor="#1E1E1E",
                   command=toggle_fields).pack(pady=5)

    account_type_var = tk.StringVar(value="Saving Account")
    tk.Label(add_frame, text="Account Type:", bg="#1E1E1E", fg="white").pack()
    account_type_menu = tk.OptionMenu(add_frame, account_type_var, "Saving Account", "Current Account")
    account_type_menu.config(bg="#3E3E3E", fg="white")
    account_type_menu.pack(pady=5)

    entry_bank = tk.Entry(add_frame, width=30)
    tk.Label(add_frame, text="Bank Name:", bg="#1E1E1E", fg="white").pack()
    entry_bank.pack(pady=5)

    entry_account = tk.Entry(add_frame, width=30)
    tk.Label(add_frame, text="Account Number:", bg="#1E1E1E", fg="white").pack()
    entry_account.pack(pady=5)

    entry_ifsc = tk.Entry(add_frame, width=30)
    tk.Label(add_frame, text="IFSC Code:", bg="#1E1E1E", fg="white").pack()
    entry_ifsc.pack(pady=5)

    entry_card = tk.Entry(add_frame, width=30)
    tk.Label(add_frame, text="Card Number:", bg="#1E1E1E", fg="white").pack()
    entry_card.pack(pady=5)

    entry_expiry = tk.Entry(add_frame, width=30)
    tk.Label(add_frame, text="Card Expiry (MM/YY):", bg="#1E1E1E", fg="white").pack()
    entry_expiry.pack(pady=5)

    entry_cvv = tk.Entry(add_frame, width=30, show="*")
    tk.Label(add_frame, text="CVV:", bg="#1E1E1E", fg="white").pack()
    entry_cvv.pack(pady=5)

    tk.Button(add_frame, text="💾 Save Banking Details", command=save_banking_data,
              bg="#28a745", fg="white", font=("Arial", 11), relief="flat").pack(pady=10)

    export_frame = tk.Frame(banking_window, bg="#1E1E1E")
    export_frame.pack(pady=10)

    tk.Button(export_frame, text="📥 Export to Excel", command=export_to_excel,
              bg="#0078D7", fg="white", font=("Arial", 10), relief="flat").pack(side="left", padx=10)
    tk.Button(export_frame, text="📄 Export to PDF", command=export_to_pdf,
              bg="#28a745", fg="white", font=("Arial", 10), relief="flat").pack(side="left", padx=10)

    tk.Button(banking_window, text="Close", command=banking_window.destroy,
              bg="#555", fg="white", font=("Arial", 11), relief="flat").pack(pady=10)

    refresh_view()

    def edit_entry(index):
        banking_data = load_banking_data()
        if index >= len(banking_data):
            return

        entry = banking_data[index]
        edit_window = tk.Toplevel()
        edit_window.title("Edit Banking Details")
        edit_window.geometry("400x500")
        edit_window.configure(bg="#1E1E1E")

        fields = {}

        # Account Type Dropdown
        tk.Label(edit_window, text="Account Type:", bg="#1E1E1E", fg="white", font=("Arial", 11)).pack(pady=5)
        account_type_var = tk.StringVar()
        account_type_var.set(entry.get("Account Type", "Saving Account"))  # Default to Saving if missing

        account_type_dropdown = tk.OptionMenu(edit_window, account_type_var, "Saving Account", "Current Account")
        account_type_dropdown.configure(bg="#3E3E3E", fg="white", highlightthickness=0)
        account_type_dropdown.pack(pady=5)

        # Other fields
        for key in ["Bank Name", "Account Number", "IFSC Code", "Card Number", "Card Expiry", "CVV"]:
            tk.Label(edit_window, text=f"{key}:", bg="#1E1E1E", fg="white", font=("Arial", 11)).pack(pady=5)
            entry_widget = tk.Entry(edit_window, width=30, bg="#3E3E3E", fg="white", insertbackground="white")
            entry_widget.insert(0, entry.get(key, ""))
            entry_widget.pack(pady=5)
            fields[key] = entry_widget

        def save_changes():
            updated_entry = {k: v.get().strip() for k, v in fields.items()}
            updated_entry["Account Type"] = account_type_var.get().strip()  # Add Account Type

            # Validation
            if not updated_entry["Account Number"].isdigit():
                show_warning_toast("Invalid Account Number", "Account Number must contain only digits.")
                return
            if not (len(updated_entry["IFSC Code"]) == 11 and updated_entry["IFSC Code"][:4].isalpha() and updated_entry["IFSC Code"][4:].isdigit()):
                show_warning_toast("Invalid IFSC", "IFSC must be 11 characters: first 4 alphabets, rest 7 digits.")
                return
            if not (updated_entry["Card Number"].isdigit() and len(updated_entry["Card Number"]) == 16):
                show_warning_toast("Invalid Card Number", "Card Number must be exactly 16 digits.")
                return
            if not re.match(r"^(0[1-9]|1[0-2])/([0-9]{2})$", updated_entry["Card Expiry"]):
                show_warning_toast("Invalid Expiry", "Card Expiry must be in MM/YY format (e.g., 01/25).")
                return
            if not (updated_entry["CVV"].isdigit() and len(updated_entry["CVV"]) == 3):
                show_warning_toast("Invalid CVV", "CVV must be exactly 3 digits.")
                return

            # Save updated
            banking_data[index] = updated_entry

            user_dir = get_user_path(current_user)
            file_path = os.path.join(user_dir, "bdata.enc")
            try:
                encrypted = cipher_suite.encrypt(str(banking_data).encode())
                with open(file_path, "wb") as f:
                    f.write(encrypted)
                show_success_toast("✅ Banking details updated successfully!")
                refresh_view()
                edit_window.destroy()
            except Exception as e:
                show_error_toast("Error: Failed to update banking data")

        tk.Button(edit_window, text="💾 Save Changes", command=save_changes, 
                  bg="#28a745", fg="white", font=("Arial", 11), relief="flat").pack(pady=10)


    def delete_entry(index):
            response = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this banking entry?")
            if not response:
                return

            banking_data = load_banking_data()
            if index >= len(banking_data):
                show_error_toast("❌ Invalid index!")
                return

            # Delete the entry
            deleted_entry = banking_data.pop(index)

            user_dir = get_user_path(current_user)
            file_path = os.path.join(user_dir, "bdata.enc")
            try:
                encrypted = cipher_suite.encrypt(str(banking_data).encode())
                with open(file_path, "wb") as f:
                    f.write(encrypted)
                show_success_toast(f"✅ Deleted entry: {deleted_entry.get('Bank Name', 'Unknown Bank')}")
                refresh_view()
            except Exception as e:
                show_error_toast(f"❌ Failed to delete banking data: {str(e)}")


            refresh_view()

            export_frame = tk.Frame(banking_window, bg="#1E1E1E")
            export_frame.pack(pady=10)

            tk.Button(export_frame, text="📥 Export to Excel", command=export_to_excel, bg="#0078D7", fg="white", font=("Arial", 10), relief="flat").pack(side="left", padx=10)
            tk.Button(export_frame, text="📄 Export to PDF", command=export_to_pdf, bg="#28a745", fg="white", font=("Arial", 10), relief="flat").pack(side="left", padx=10)

            tk.Button(banking_window, text="Close", command=banking_window.destroy, bg="#555", fg="white", font=("Arial", 11), relief="flat").pack(pady=10)
            refresh_view()
def load_bank_data():
        user_dir = get_user_path(current_user)
        file_path = os.path.join(user_dir, "bdata.enc")
        if not os.path.exists(file_path):
            return []
        try:
            with open(file_path, "rb") as f:
                decrypted = cipher_suite.decrypt(f.read()).decode()
            data = eval(decrypted)
            if isinstance(data, dict):
                return [data]
            return data
        except:
            return []

def get_banking_entry_count():
    try:
        banking_data = load_bank_data()
        return len(banking_data) if banking_data else 0
    except Exception as e:
        show_error_toast(f"Error getting banking data: {e}")
        return 0




import re

def save_personal_ids():
    global current_user
    if not current_user:
        show_error_toast("❌ Error: No logged-in user!")
        return

    user_dir = get_user_path(current_user)
    id_file = os.path.join(user_dir, "personal_ids.enc")

    fields = {
        "Aadhaar Card Number": "",
        "PAN Card Number": "",
        "Driving License Number": "",
        "Vehicle Registration Number": "",
    }

    custom_ids = {}

    id_window = tk.Toplevel()
    id_window.title("Save Personal IDs")
    id_window.geometry("450x700")
    id_window.configure(bg="#2E2E2E")

    entries = {}

    try:
        if os.path.exists(id_file):
            with open(id_file, "r") as f:
                existing_data = json.load(f)
        else:
            existing_data = {}
    except Exception as e:
        show_error_toast(f"❌ Error reading existing IDs: {str(e)}")
        existing_data = {}

    def validate_fields(data):
        aadhaar = data.get("Aadhaar Card Number", "")
        if aadhaar and not re.fullmatch(r"\d{12}", aadhaar):
            show_warning_toast("⚠️ Invalid Aadhaar Number! (12 digits required)")
            return False

        pan = data.get("PAN Card Number", "")
        if pan and not re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", pan, re.IGNORECASE):
            show_warning_toast("⚠️ Invalid PAN Number! (Format: ABCDE1234F)")
            return False

        dl = data.get("Driving License Number", "")
        if dl and not re.fullmatch(r"[A-Z]{2}\d{2}\s?\d{11}", dl, re.IGNORECASE):
            show_warning_toast("⚠️ Invalid Driving License Number!")
            return False

        reg = data.get("Vehicle Registration Number", "")
        if reg and not re.fullmatch(r"[A-Z]{2}\d{2}[A-Z]{1,2}\d{4}", reg, re.IGNORECASE):
            show_warning_toast("⚠️ Invalid Vehicle Registration Number!")
            return False

        for name, value in custom_ids.items():
            if not re.fullmatch(r"[\w\s\-\/]{3,50}", name):
                show_warning_toast(f"⚠️ Invalid Custom ID Name: '{name}'")
                return False
            if not re.fullmatch(r"[\w\-\/]{3,50}", value):
                show_warning_toast(f"⚠️ Invalid Custom ID Value for '{name}'")
                return False

        return True

    def force_uppercase(event):
        content = event.widget.get()
        event.widget.delete(0, tk.END)
        event.widget.insert(0, content.upper())
        event.widget.icursor(tk.END)

    def add_custom_id():
        id_name = custom_type_entry.get().strip()
        id_value = custom_value_entry.get().strip()

        if not id_name or not id_value:
            show_warning_toast("⚠️ Enter both ID Name and ID Number!")
            return

        all_values = list(custom_ids.values()) + [entry.get().strip() for entry in entries.values()]
        if id_value in all_values:
            show_warning_toast("⚠️ ID Number must be unique!")
            return

        custom_ids[id_name] = id_value
        custom_listbox.insert(tk.END, f"{id_name}: {id_value}")
        custom_type_entry.delete(0, tk.END)
        custom_value_entry.delete(0, tk.END)
        show_success_toast("✅ Custom ID Added!")

    def delete_selected_custom_id():
        selected_index = custom_listbox.curselection()
        if not selected_index:
            show_warning_toast("⚠️ Select a Custom ID to delete!")
            return

        selected_item = custom_listbox.get(selected_index)
        id_name = selected_item.split(":")[0].strip()
        if id_name in custom_ids:
            del custom_ids[id_name]
        custom_listbox.delete(selected_index)
        show_success_toast("✅ Custom ID Deleted!")

    def save_ids_to_file():
        updated_data = existing_data.copy()

        for label, entry in entries.items():
            try:
                value = entry.get().strip()
                updated_data[label] = value
            except Exception as e:
                show_error_toast(f"❌ Entry error: {e}")

        updated_data.update(custom_ids)

        if not validate_fields(updated_data):
            return

        with open(id_file, "w") as f:
            json.dump(updated_data, f, indent=4)
        show_success_toast("✅ Personal IDs Saved Successfully!")
        id_window.destroy()

    for label in fields:
        tk.Label(id_window, text=label, fg="white", bg="#2E2E2E", font=("Arial", 11)).pack(pady=5)
        entry = tk.Entry(id_window, width=40, bg="#3E3E3E", fg="white", insertbackground="white")
        value = existing_data.get(label, "")
        if isinstance(value, str):
            entry.insert(0, value)
        if label in ["PAN Card Number", "Driving License Number", "Vehicle Registration Number"]:
            entry.bind("<KeyRelease>", force_uppercase)
        entry.pack(pady=5)
        entries[label] = entry

    tk.Label(id_window, text="Custom ID Name", fg="white", bg="#2E2E2E", font=("Arial", 11)).pack(pady=5)
    custom_type_entry = tk.Entry(id_window, width=40, bg="#3E3E3E", fg="white", insertbackground="white")
    custom_type_entry.pack(pady=5)

    tk.Label(id_window, text="Custom ID Number", fg="white", bg="#2E2E2E", font=("Arial", 11)).pack(pady=5)
    custom_value_entry = tk.Entry(id_window, width=40, bg="#3E3E3E", fg="white", insertbackground="white")
    custom_value_entry.pack(pady=5)

    tk.Button(id_window, text="Add Custom ID", command=add_custom_id, bg="#0078D7", fg="white",
              font=("Arial", 11), relief="flat").pack(pady=10)

    custom_listbox = tk.Listbox(id_window, width=45, height=6, bg="#3E3E3E", fg="white")
    custom_listbox.pack(pady=10)

    for k, v in existing_data.items():
        if k not in fields:
            custom_ids[k] = v
            custom_listbox.insert(tk.END, f"{k}: {v}")

    tk.Button(id_window, text="Delete Selected Custom ID", command=delete_selected_custom_id,
              bg="#dc3545", fg="white", font=("Arial", 11), relief="flat").pack(pady=5)

    tk.Button(id_window, text="Save All IDs", command=save_ids_to_file, bg="#4CAF50", fg="white",
              font=("Arial", 12), relief="flat").pack(pady=20)

    id_window.mainloop()



def view_personal_ids():
    global current_user
    if not current_user:
        show_error_toast("❌ Error: No logged-in user!")
        return

    user_dir = get_user_path(current_user)
    id_file = os.path.join(user_dir, "personal_ids.enc")

    if not os.path.exists(id_file):
        show_error_toast("❌ Error: No Personal IDs saved yet!")
        return

    try:
        with open(id_file, "r") as f:
            saved_data = json.load(f)

        view_window = tk.Toplevel()
        view_window.title("View & Manage Personal IDs")
        view_window.geometry("500x600")
        view_window.configure(bg="#2E2E2E")

        canvas = tk.Canvas(view_window, bg="#2E2E2E", highlightthickness=0)
        scrollbar = tk.Scrollbar(view_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#2E2E2E")

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        entry_widgets = {}

        def enable_edit(entry):
            entry.config(state="normal")

        def save_changes():
            updated_data = {}
            id_numbers_seen = set()

            for key, (label, entry, _) in entry_widgets.items():
                value = entry.get().strip()
                if not value:
                    continue

                if value in id_numbers_seen:
                    show_warning_toast(f"⚠️ Duplicate ID Number found: {value}")
                    return

                id_numbers_seen.add(value)
                updated_data[key] = value

            with open(id_file, "w") as f:
                json.dump(updated_data, f, indent=4)

            show_success_toast("✅ Changes Saved Successfully!")
            view_window.destroy()

        def delete_id(key):
            if key in entry_widgets:
                label_widget, entry_widget, edit_btn = entry_widgets[key]
                label_widget.destroy()
                entry_widget.destroy()
                edit_btn.destroy()
                del entry_widgets[key]
                show_success_toast(f"✅ '{key}' deleted.")

        for key, value in saved_data.items():
            label = tk.Label(scrollable_frame, text=f"{key}:", fg="white", bg="#2E2E2E", font=("Arial", 11, "bold"))
            label.pack(anchor="w", padx=10, pady=(10, 2))

            frame = tk.Frame(scrollable_frame, bg="#2E2E2E")
            frame.pack(fill="x", padx=20)

            entry = tk.Entry(frame, width=30, bg="#3E3E3E", fg="white", insertbackground="white")
            entry.insert(0, value)
            entry.config(state="disabled")
            entry.pack(side="left", padx=(0, 5), pady=2)

            edit_btn = tk.Button(frame, text="Edit", command=lambda e=entry: enable_edit(e),
                                 bg="#0078D7", fg="white", relief="flat", font=("Arial", 9))
            edit_btn.pack(side="right", padx=5)

            delete_btn = tk.Button(frame, text="Delete", command=lambda k=key: delete_id(k),
                                   bg="#dc3545", fg="white", relief="flat", font=("Arial", 9))
            delete_btn.pack(side="right", padx=5)

            entry_widgets[key] = (label, entry, edit_btn)

        tk.Button(view_window, text="Save All Changes", command=save_changes,
                  bg="#4CAF50", fg="white", font=("Arial", 12), relief="flat").pack(pady=20)

    except Exception as e:
        show_error_toast(f"❌ Error loading Personal IDs: {str(e)}")




        
def main_app():
    global root, website_entry, username_entry, password_entry, stats_label
    
    root = tk.Tk()
    root.title("Password Manager 5.1")
    root.geometry("750x700")
    root.configure(bg="#1E1E1E")
    root.attributes('-alpha', 0.9)
    root.resizable(True, True)
    root.protocol("WM_DELETE_WINDOW", on_exit)

    menubar = tk.Menu(root, bg="#1E1E1E", fg="white", activebackground="#333333", relief="flat", font=("Arial", 12, "bold"), bd=3)

    def themed_menu():
        return tk.Menu(menubar, tearoff=0, bg="#1E1E1E", fg="white", activebackground="#333333", activeforeground="white", font=("Arial", 12), bd=3)

    # Menus
    home_menu = themed_menu(); home_menu.add_command(label="Home", command=lambda: show_tab("home"))
    about_menu = themed_menu(); about_menu.add_command(label="About", command=lambda: show_tab("about"))
    password_menu = themed_menu()
    password_menu.add_command(label="Generate Password", command=show_generated_password)
    password_menu.add_separator()
    password_menu.add_command(label="Retrieve Passwords", command=retrieve_password)
    password_menu.add_separator()
    password_menu.add_command(label="Delete Passwords", command=delete_password)
    password_menu.add_separator()
    password_menu.add_command(label="Import from Excel", command=import_from_excel)
    password_menu.add_separator()
    password_menu.add_command(label="Export to Excel", command=export_passwords_to_excel)
    password_menu.add_separator()
    password_menu.add_command(label="Password Health Dashboard", command=show_password_health_dashboard)
    encryption_menu = themed_menu(); encryption_menu.add_command(label="Encrypt Files", command=encrypt_files)
    encryption_menu.add_separator()
    encryption_menu.add_command(label="Decrypt File", command=decrypt_file)
    encryption_menu.add_separator()
    encryption_menu.add_command(label="Decrypt All File", command=decrypt_all_files)
    encryption_menu.add_separator()
    encryption_menu.add_command(label="View Encrypted Files", command=view_encrypted_files)
    encryption_menu.add_separator()
    encryption_menu.add_command(label="🖼 View Encrypted Image/Video", command=decrypt_and_view_image)
     
    notes_menu = themed_menu(); notes_menu.add_command(label="Create Note", command=create_note)
    notes_menu.add_separator()
    notes_menu.add_command(label="View Notes", command=view_notes)
    notes_menu.add_separator()
    notes_menu.add_command(label="Export Notes", command=export_notes)    
    other_menu = themed_menu(); other_menu.add_command(label="Banking", command=launch_banking_ui)
    other_menu.add_separator()
    other_menu.add_command(label="Add Personal ID's", command=save_personal_ids)
    other_menu.add_separator()
    other_menu.add_command(label="View Personal ID's", command=view_personal_ids)
    account_menu = themed_menu(); account_menu.add_command(label="Delete User Account", command=delete_user)
    account_menu.add_separator()
    account_menu.add_command(label="Backup ", command=backup_current_user_data)
    account_menu.add_separator()
    account_menu.add_command(label="Restore", command=restore_current_user_data)
    account_menu.add_separator()
    account_menu.add_command(label="Forget PIN", command=forget_pin)
    account_menu.add_separator()
    account_menu.add_command(label="Logout", command=logout)
    menubar.add_cascade(label="Home", menu=home_menu)
    menubar.add_cascade(label="Passwords", menu=password_menu)
    menubar.add_cascade(label="Files", menu=encryption_menu)
    menubar.add_cascade(label="Notes", menu=notes_menu)
    menubar.add_cascade(label="Other", menu=other_menu)
    menubar.add_cascade(label="Account", menu=account_menu)
    menubar.add_cascade(label="About", menu=about_menu)

    root.config(menu=menubar)

    tab_content_frame = tk.Frame(root, bg="#1E1E1E")
    tab_content_frame.pack(fill="both", expand=True)

    def refresh_stats():
        if stats_label.winfo_exists():
            stats_label.config(text=get_summary_stats())
        root.after(5000, refresh_stats)

    def show_tab(tab_name):
        for widget in tab_content_frame.winfo_children():
            widget.destroy()

        if tab_name == "home":
            tk.Label(tab_content_frame, text=f"Welcome {current_user}", font=("Segoe UI", 16, "bold"), bg="#1E1E1E", fg="white").pack(pady=20)

            global stats_label
            stats_label = tk.Label(tab_content_frame, text=get_summary_stats(), font=("Segoe UI", 13), bg="#1E1E1E", fg="white", justify="left")
            stats_label.pack(pady=10, anchor="w", padx=20)
            refresh_stats()

            if is_trial_version():
                usage_count = load_license_data().get("trial_uses", 0)
                remaining = 10 - usage_count

                top_right_frame = tk.Frame(tab_content_frame, bg="#1E1E1E")
                top_right_frame.place(relx=1.0, rely=0.0, x=-20, y=10, anchor="ne")

                tk.Label(top_right_frame, text=f"Trial Mode: {remaining} use(s) left", font=("Segoe UI", 12), fg="orange", bg="#1E1E1E").pack()
                tk.Button(top_right_frame, text="Upgrade License", command=upgrade,
                          bg="#FFA500", fg="black", font=("Segoe UI", 12, "bold"), relief="flat").pack(pady=5)

            global website_entry, username_entry, password_entry
            field_font = ("Segoe UI", 14)
            entry_style = {
                "font": field_font,
                "bg": "#2E2E2E",
                "fg": "white",
                "insertbackground": "white",
                "width": 30,
                "relief": "flat",
                "highlightthickness": 1,
                "highlightbackground": "#555"
            }

            for field_name, var in [("Website", "website_entry"), ("Username", "username_entry"), ("Password", "password_entry")]:
                tk.Label(tab_content_frame, text=field_name, bg="#1E1E1E", fg="white", font=field_font).pack(pady=(20, 5))
                entry = tk.Entry(tab_content_frame, **entry_style, show="*" if field_name == "Password" else "")
                entry.pack(pady=5)
                globals()[var] = entry

            tk.Button(tab_content_frame, text="Save Password", command=save_password,
                      bg="#4B0082", fg="white", font=("Segoe UI", 14, "bold"),
                      activebackground="#9932CC", relief="flat", width=20, height=2).pack(pady=30)


        elif tab_name == "about":
            license_data = load_license_data()
            license_key = license_data.get("license_key", "Not activated")

            if license_key == "T-050505ABCDE":
                license_status = "Trial Mode"
                trial_uses = license_data.get("trial_uses", 0)
                uses_left = max(0, 10 - trial_uses)
            elif license_key in load_license_keys():
                license_status = "Permanent Activated"
                uses_left = None
            else:
                license_status = "Not Activated"
                uses_left = None

            masked_key = f"{license_key[:3]}****{license_key[-4:]}" if license_key != "Not activated" else license_key

            tk.Label(tab_content_frame, text="🔐 Password Manager", font=("Segoe UI", 18, "bold"), fg="white", bg="#1E1E1E").pack(pady=10)
            tk.Label(tab_content_frame, text="Version: 5.1", font=("Segoe UI", 13), fg="gray", bg="#1E1E1E").pack(pady=5)
            tk.Label(tab_content_frame, text=f"License Key: {masked_key}", font=("Segoe UI", 13), fg="orange", bg="#1E1E1E").pack(pady=5)
            tk.Label(tab_content_frame, text=f"Status: {license_status}", font=("Segoe UI", 13), fg="lightgreen" if license_status == "Permanent Activated" else "orange", bg="#1E1E1E").pack(pady=5)

            if uses_left is not None:
                tk.Label(tab_content_frame, text=f"Trial Uses Left: {uses_left}/10", font=("Segoe UI", 13), fg="yellow", bg="#1E1E1E").pack(pady=5)

            tk.Label(tab_content_frame, text="Created By: Vikash Agnihotry", font=("Segoe UI", 13), fg="gray", bg="#1E1E1E").pack(pady=10)

            if is_trial_version():
                tk.Button(tab_content_frame, text="Upgrade License", command=upgrade_to_permanent,
                          bg="#FFA500", fg="black", font=("Segoe UI", 12, "bold"), relief="flat").pack(pady=10)



    show_tab("home")
    root.mainloop()


    
    # Right Panel (Content Frame)
def get_summary_stats():
    try:
        if not current_user:
            return "❌ No user logged in."

        username = current_user
        user_dir = get_user_path(username)
        pass_file = os.path.join(user_dir, "pass.enc")
        encrypted_list_file = os.path.join(user_dir, "list.enc")
        notes_file = os.path.join(user_dir, "notes.enc")
        id_file = os.path.join(user_dir, "personal_ids.enc")  # Now using JSON instead of ENC

        # Count passwords
        password_data = decrypt_data(pass_file) if os.path.exists(pass_file) else {}
        total_passwords = sum(len(v) if isinstance(v, list) else 1 for v in password_data.values())

        # Count encrypted files
        encrypted_files = decrypt_data(encrypted_list_file) if os.path.exists(encrypted_list_file) else []
        total_encrypted = len(encrypted_files)

        # Count notes
        total_notes = 0
        if os.path.exists(notes_file):
            with open(notes_file, "r") as f:
                notes = [line for line in f if line.strip()]  # Ignore empty lines
            total_notes = len(notes)

        # Banking entries
        total_banking = get_banking_entry_count() if 'get_banking_entry_count' in globals() else 0

        # Personal IDs count
        if os.path.exists(id_file):
            with open(id_file, "r") as f:
                id_data = json.load(f)
            total_ids = sum(1 for v in id_data.values() if v.strip())

        else:
            total_ids = 0

        return (
            f"🔐 Total Passwords Stored: {total_passwords}\n"
            f"📁 Total Encrypted Files: {total_encrypted}\n"
            f"📝 Total Notes Saved: {total_notes}\n"
            f"🏦 Banking Details Stored: {total_banking}\n"
            f"🔖 Total Personal IDs Stored: {total_ids}"
        )

    except Exception as e:
        return f"❌ Error retrieving summary: {e}"


def save_entry():
    # Create a new top-level window (secondary window)
    save_window = tk.Toplevel(root)
    save_window.title("Save New Password")
    save_window.geometry("400x400")  # Adjust the size of the new window as needed
    save_window.configure(bg='#1E1E1E')

    # Website/Service Label and Entry
    tk.Label(save_window, text="Website/Service", fg='white', bg='#1E1E1E', font=("Arial", 12)).pack(pady=5)
    website_entry = tk.Entry(save_window)
    website_entry.pack(fill="x", padx=10, pady=5)

    # Username Label and Entry
    tk.Label(save_window, text="Username", fg='white', bg='#1E1E1E', font=("Arial", 12)).pack(pady=5)
    username_entry = tk.Entry(save_window)
    username_entry.pack(fill="x", padx=10, pady=5)

    # Password Label and Entry
    tk.Label(save_window, text="Password", fg='white', bg='#1E1E1E', font=("Arial", 12)).pack(pady=5)
    password_entry = tk.Entry(save_window, show="*")
    password_entry.pack(fill="x", padx=10, pady=5)

    # Save Button
    tk.Button(save_window, text="Save", command=save_password, 
              bg='#4B0082', fg='white', font=("Arial", 12), relief='flat', 
              activebackground='#218838', width=20).pack(pady=10)
    
    
def save_password():
    user_dir = get_user_path(current_user)
    pass_file = os.path.join(user_dir, "pass.enc")
    website, username, password = website_entry.get(), username_entry.get(), password_entry.get()

    if not (website and username and password):
        show_warning_toast("Please fill all fields.")
        return

    encrypted_password = cipher_suite.encrypt(password.encode()).decode()
    data = decrypt_data(pass_file) or {}  # Ensure data is always a dictionary

    # Check if the exact same entry exists (skip saving)
    for existing_website, entries in data.items():
        if isinstance(entries, dict):  # Convert old format
            data[existing_website] = [entries]

        if existing_website == website:
            for entry in data[existing_website]:
                if entry["username"] == username:
                    if entry["password"] == encrypted_password:
                        show_error_toast("❌ Error: User Name & Password Already Exists ")
                        return  # Skip saving
                    else:
                        update = messagebox.askyesno(
                            "Password Update",
                            f"Username '{username}' for '{website}' exists with a different password. Update?"
                        )
                        if update:
                            entry["password"] = encrypted_password  # Update existing password
                            encrypt_data(data, pass_file)  # Save updated data
                            show_success_toast("✅Password Updated Successfully")
                        else:
                             show_warning_toast("⚠Password Update Skipped!")
                        return  # Exit function

    # If website exists but username is different, create a new separate website entry
    unique_website = website
    count = 1
    while unique_website in data:
        unique_website = f"{website} ({count})"  # Rename website to create separate entry
        count += 1

    data[unique_website] = [{"username": username, "password": encrypted_password}]

    encrypt_data(data, pass_file)  # Save updated data

    show_success_toast("✅ New Password Saved Successfully!")

    # Clear input fields only on successful save
    website_entry.delete(0, tk.END)
    username_entry.delete(0, tk.END)
    password_entry.delete(0, tk.END)


def retrieve_password():
    user_dir = get_user_path(current_user)
    pass_file = os.path.join(user_dir, "pass.enc")
    
    data = decrypt_data(pass_file)
    if not data:
        show_error_toast("❌ Error: No Stored Password Found ")
        return
    
    display_all_passwords(data)

def display_all_passwords(data):
    top = Toplevel()
    top.title("Stored Passwords")
    top.geometry("600x500")
    top.configure(bg="#222")
    
    frame = Frame(top, bg="#222")
    frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
    
    scrollbar = Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    listbox = Listbox(frame, yscrollcommand=scrollbar.set, width=80, height=20, bg="#333", fg="white", font=("Arial", 12), relief="flat")
    for website, accounts in data.items():
        if isinstance(accounts, dict):
            accounts = [accounts]
        listbox.insert(tk.END, f"Website: {website}")
        for account in accounts:
            decrypted_password = cipher_suite.decrypt(account["password"].encode()).decode()
            listbox.insert(tk.END, f"  Username: {account['username']}")
            listbox.insert(tk.END, f"  Password: {decrypted_password}")
            listbox.insert(tk.END, "-"*40)
    
    listbox.pack(fill=tk.BOTH, expand=True)
    scrollbar.config(command=listbox.yview)
        
def display_passwords(accounts):
    top = Toplevel()
    top.title("Stored Accounts")
    top.geometry("400x300")
    
    scrollbar = Scrollbar(top)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    listbox = Listbox(top, yscrollcommand=scrollbar.set)
    for account in accounts:
        decrypted_password = cipher_suite.decrypt(account["password"].encode()).decode()
        listbox.insert(tk.END, f"Username: {account['username']} | Password: {decrypted_password}")
    
    listbox.pack(fill=tk.BOTH, expand=True)
    scrollbar.config(command=listbox.yview)
def delete_password():
    user_dir = get_user_path(current_user)
    pass_file = os.path.join(user_dir, "pass.enc")
    data = decrypt_data(pass_file) or {}  # Ensure we get a dictionary, not None

    if not data:
        show_error_toast("❌ Error: No Stored Password Found ")
        return

    # Create delete window
    delete_window = create_window("Delete Passwords", "600x400")

    # Scrollable listbox
    frame = Frame(delete_window, bg='#333')
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    listbox = Listbox(frame, selectmode="multiple", bg='#2E2E2E', fg='white', font=("Arial", 12))
    listbox.pack(side="left", fill="both", expand=True)

    scrollbar = Scrollbar(frame, orient="vertical", command=listbox.yview)
    scrollbar.pack(side="right", fill="y")
    listbox.config(yscrollcommand=scrollbar.set)

    # Populate the listbox
    password_entries = []  # Store (website, index) mapping for deletion

    for website, credentials_list in data.items():
        for index, credentials in enumerate(credentials_list):
            display_text = f"{website} - {credentials['username']}"
            password_entries.append((website, index))  # Store corresponding data
            listbox.insert("end", display_text)

    def confirm_delete():
        selected_indices = listbox.curselection()
        if not selected_indices:
            show_warning_toast("⚠Select Atleast 1 Entry to Delete!")
            return

        updated_data = {}
        passwords_deleted = False

        for i, (website, index) in enumerate(password_entries):
            if i in selected_indices:
                passwords_deleted = True  # Mark that we are deleting this entry
                continue  # Skip adding this entry

            if website not in updated_data:
                updated_data[website] = []
            updated_data[website].append(data[website][index])

        encrypt_data(updated_data, pass_file)  # Save updated passwords
        show_success_toast("✅Password Successfully Deleted!")
        delete_window.destroy()

    # Delete button
    Button(
        delete_window, text="Delete Selected", command=confirm_delete, bg='#D9534F', fg='white',
        font=("Arial", 12), relief='flat', activebackground='#C9302C'
    ).pack(pady=10)






root = Tk()
root.withdraw()  # Hide root until needed

if check_license():
    check_trail_license()  # Corrected from 'check_trail_license'
    delete_status_file()
    
    # Now check if the user directory is empty
    if not any(os.scandir(DATA_DIR)):
        generate_all_fakes()
        register()
    else:
        open_login_window()
else:
    create_fake_system_dlls()
    activate_license()

root.mainloop()
